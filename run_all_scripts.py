from __future__ import annotations

import argparse
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import conciliador_planilhas_sku as conciliador
from msn_utils import emit, json_safe, setup_script_logging, write_json

DEFAULT_CONCILIACAO_FOLDER = Path.home() / "Desktop" / "Conciliacao"
DEFAULT_DESKTOP = Path.home() / "Desktop"

SCRIPT_ROOT = Path(__file__).resolve().parent
CONCILIADOR = SCRIPT_ROOT / "conciliador_planilhas_sku.py"
BUSCADOR = SCRIPT_ROOT / "buscador_candidatas_imagens.py"
OTIMIZADOR = SCRIPT_ROOT / "otimizador_imagens.py"
IMPORTADOR_WOO = SCRIPT_ROOT / "importador_piloto_woocommerce.py"
VERIFICADOR = SCRIPT_ROOT / "verificador_fluxo_msn.py"


def run_script(command: list[str], logger: object | None = None, display_command: list[str] | None = None) -> int:
    emit(logger, "\n=> Executando: " + " ".join(display_command or command))
    completed = subprocess.run(command, text=True)
    if completed.returncode != 0:
        emit(logger, f"Erro: comando retornou {completed.returncode}", error=True)
    return completed.returncode


def expected_conciliador_output(args: argparse.Namespace, conciliacao_folder: Path) -> Path:
    if not args.saida:
        return conciliacao_folder / "todos-os-produtos.xlsx"

    if args.wordpress:
        wordpress_path = args.wordpress.expanduser().resolve()
    else:
        wordpress_path = conciliador.find_xlsx_by_keyword(conciliacao_folder, "wordpress")
    return conciliador.output_path(wordpress_path, args.saida, fallback_suffix="_atualizada")


def expected_relatorio_output(args: argparse.Namespace, todos_workbook: Path, conciliacao_folder: Path) -> Path:
    if args.relatorio:
        return conciliador.report_path(todos_workbook, args.relatorio)
    return conciliacao_folder / "relatorio-conciliacao.xlsx"


def expected_novos_output(args: argparse.Namespace, todos_workbook: Path, conciliacao_folder: Path) -> Path:
    if args.saida_novos_produtos:
        return conciliador.output_path(todos_workbook, args.saida_novos_produtos, fallback_suffix="_novos")
    return conciliacao_folder / "produtos-novos.xlsx"


def expected_woo_pilot_workbook(args: argparse.Namespace, novos_workbook: Path, conciliacao_folder: Path) -> Path:
    if args.woo_workbook:
        return args.woo_workbook.expanduser().resolve()
    sample = conciliacao_folder / "produtos-novos-amostra-5.xlsx"
    if sample.exists():
        return sample
    return novos_workbook


def expected_verification_output(conciliacao_folder: Path) -> Path:
    return conciliacao_folder / "verificacao-fluxo-msn.json"


def write_manifest(path: Path | None, manifest: dict[str, object]) -> None:
    if path is None:
        return
    write_json(path, manifest)


def summarize_conciliacao_report(path: Path) -> dict[str, object]:
    if not path.exists():
        return {}
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook["conciliacao"] if "conciliacao" in workbook.sheetnames else workbook.active
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        try:
            status_index = headers.index("Status_Conciliacao") + 1
        except ValueError:
            workbook.close()
            return {}
        counts: dict[str, int] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            status = str(row[status_index - 1] or "").strip() or "vazio"
            counts[status] = counts.get(status, 0) + 1
        workbook.close()
        return {
            "status_counts": counts,
            "atualizados": counts.get("atualizado_por_nome", 0) + counts.get("atualizado_por_sku", 0),
            "novos": counts.get("adicionado_ao_wordpress", 0),
            "ignorados": counts.get("ignorado_sem_nome", 0),
            "erros": counts.get("erro_validacao", 0),
        }
    except Exception as exc:
        return {"summary_error": str(exc)}


def manifest_args(args: argparse.Namespace, conciliacao_folder: Path, manifesto_path: Path, log_path: Path) -> dict[str, object]:
    raw_args = vars(args).copy()
    for secret_key in ("woo_consumer_key", "woo_consumer_secret"):
        if raw_args.get(secret_key):
            raw_args[secret_key] = "***"
    raw_args.update(
        {
            "conciliacao_folder": str(conciliacao_folder),
            "desktop_folder_name": args.desktop_folder_name,
            "manifesto": str(manifesto_path),
            "log": str(log_path),
        }
    )
    return json_safe(raw_args)  # type: ignore[return-value]


def build_conciliador_command(args: argparse.Namespace, conciliacao_folder: Path) -> list[str]:
    command = [
        sys.executable,
        str(CONCILIADOR),
        "--conciliacao-folder",
        str(conciliacao_folder),
    ]
    optional_pairs = [
        ("--sheet-cliente", args.sheet_cliente),
        ("--sheet-wordpress", args.sheet_wordpress),
        ("--cliente", args.cliente),
        ("--wordpress", args.wordpress),
        ("--saida", args.saida),
        ("--relatorio", args.relatorio),
        ("--saida-novos-produtos", args.saida_novos_produtos),
    ]
    if args.proximo_id is not None:
        command.extend(["--proximo-id", str(args.proximo_id)])
    append_optional_pairs(command, optional_pairs)
    return command


def build_verifier_command(args: argparse.Namespace, conciliacao_folder: Path, verification_output: Path) -> list[str]:
    command = [
        sys.executable,
        str(VERIFICADOR),
        "--conciliacao-folder",
        str(conciliacao_folder),
        "--summary-json",
        str(verification_output),
    ]
    if args.log_dir:
        command.extend(["--log-dir", str(args.log_dir)])
    return command


def build_woo_command(args: argparse.Namespace, woo_pilot_workbook: Path) -> tuple[list[str], list[str]]:
    command = [
        sys.executable,
        str(IMPORTADOR_WOO),
        "--workbook",
        str(woo_pilot_workbook),
        "--limit",
        str(args.woo_limit),
        "--status",
        args.woo_status,
    ]
    flags = [
        ("--apply", args.woo_apply),
        ("--preflight-only", args.woo_preflight_only),
        ("--update-existing", args.woo_update_existing),
    ]
    for flag, enabled in flags:
        if enabled:
            command.append(flag)
    append_optional_pairs(
        command,
        [
            ("--env-file", args.woo_env_file),
            ("--site-url", args.woo_site_url),
            ("--consumer-key", args.woo_consumer_key),
            ("--consumer-secret", args.woo_consumer_secret),
            ("--timeout", args.timeout),
            ("--retries", args.retries),
            ("--retry-backoff", args.retry_backoff),
            ("--log-dir", args.log_dir),
        ],
    )
    return command, redact_command(command, {"--consumer-key", "--consumer-secret"})


def build_buscador_command(args: argparse.Namespace, conciliacao_folder: Path, todos_workbook: Path, candidate_root: Path) -> list[str]:
    command = [
        sys.executable,
        str(BUSCADOR),
        "--source-root",
        str(conciliacao_folder),
        "--workbook",
        str(todos_workbook),
        "--web",
    ]
    if not args.no_download:
        command.extend(["--download", "--download-root", str(candidate_root)])
    if args.include_existing_products:
        command.append("--include-existing-products")
    append_optional_pairs(
        command,
        [
            ("--sheet", args.search_sheet),
            ("--timeout", args.timeout),
            ("--retries", args.retries),
            ("--retry-backoff", args.retry_backoff),
        ],
    )
    return command


def build_otimizador_command(args: argparse.Namespace, candidate_root: Path, todos_workbook: Path) -> list[str]:
    command = [
        sys.executable,
        str(OTIMIZADOR),
        "--downloads-dir",
        str(candidate_root),
        "--workbook",
        str(todos_workbook),
    ]
    if args.white_background:
        command.append("--white-background")
    if args.overwrite:
        command.append("--overwrite")
    append_optional_pairs(
        command,
        [
            ("--timeout", args.timeout),
            ("--download-timeout", args.download_timeout),
            ("--retries", args.retries),
            ("--retry-backoff", args.retry_backoff),
        ],
    )
    return command


def append_optional_pairs(command: list[str], pairs: list[tuple[str, object | None]]) -> None:
    for flag, value in pairs:
        if value is not None:
            command.extend([flag, str(value)])


def redact_command(command: list[str], secret_flags: set[str]) -> list[str]:
    return ["***" if index > 0 and command[index - 1] in secret_flags else value for index, value in enumerate(command)]


def record_step(manifest: dict[str, object], name: str, command: list[str], exit_code: int) -> None:
    manifest["steps"].append({"name": name, "command": command, "exit_code": exit_code})  # type: ignore[index]


def finish_manifest(manifest: dict[str, object], manifesto_path: Path, exit_code: int) -> int:
    manifest["finished_at"] = datetime.now().isoformat(timespec="seconds")
    manifest["exit_code"] = exit_code
    write_manifest(manifesto_path, manifest)
    return exit_code


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Executa os scripts MSN em sequencia local.")
    parser.add_argument(
        "--conciliacao-folder",
        type=Path,
        default=DEFAULT_CONCILIACAO_FOLDER,
        help="Pasta Desktop/Conciliacao onde estao os arquivos cliente e wordpress.",
    )
    parser.add_argument(
        "--desktop-folder-name",
        required=True,
        help="Nome da pasta no Desktop onde o buscador criará as subpastas por SKU.",
    )
    parser.add_argument("--cliente", type=Path, help="Caminho da planilha do cliente para conciliador.")
    parser.add_argument("--wordpress", type=Path, help="Caminho da planilha do WordPress para conciliador.")
    parser.add_argument("--saida", type=Path, help="Pasta ou arquivo de saida para o conciliador. Recomendado usar default no Conciliacao.")
    parser.add_argument("--relatorio", type=Path, help="Pasta ou arquivo de relatorio para o conciliador.")
    parser.add_argument(
        "--saida-novos-produtos",
        type=Path,
        help="Pasta ou arquivo de saida para o arquivo de novos produtos do conciliador.",
    )
    parser.add_argument("--all", action="store_true", help="Executa todos os passos: conciliador, buscador e otimizador.")
    parser.add_argument("--sheet-cliente", help="Aba da planilha do cliente.")
    parser.add_argument("--sheet-wordpress", help="Aba da planilha WordPress.")
    parser.add_argument(
        "--proximo-id",
        type=int,
        help="Opcao legada; novos produtos ficam sem ID para o WooCommerce gerar IDs pelo SKU.",
    )
    parser.add_argument("--search-sheet", help="Aba da planilha para busca de imagens.")
    parser.add_argument(
        "--include-existing-products",
        action="store_true",
        help="Faz o buscador processar todos os produtos, mesmo os que já têm imagens locais.",
    )
    parser.add_argument("--overwrite", action="store_true", help="Força reprocessar imagens existentes no otimizador.")
    parser.add_argument("--white-background", action="store_true", help="Gera WebP com fundo branco no otimizador.")
    parser.add_argument("--no-download", action="store_true", help="Nao baixa imagens durante a busca de candidatas.")
    parser.add_argument("--timeout", type=float, help="Timeout HTTP repassado ao buscador/otimizador.")
    parser.add_argument("--download-timeout", type=float, help="Timeout HTTP de downloads repassado ao otimizador.")
    parser.add_argument("--retries", type=int, help="Retentativas HTTP repassadas ao buscador/otimizador.")
    parser.add_argument("--retry-backoff", type=float, help="Backoff HTTP repassado ao buscador/otimizador.")
    parser.add_argument("--skip-images", action="store_true", help="Roda apenas o conciliador e para antes do buscador/otimizador.")
    parser.add_argument("--skip-verify", action="store_true", help="Nao roda o verificador local apos o conciliador.")
    parser.add_argument("--woo-pilot", action="store_true", help="Roda importacao piloto WooCommerce em dry-run apos o conciliador.")
    parser.add_argument("--woo-apply", action="store_true", help="Com --woo-pilot, executa criacao real no WooCommerce.")
    parser.add_argument("--woo-preflight-only", action="store_true", help="Com --woo-pilot, valida API WooCommerce e nao cria produtos.")
    parser.add_argument("--woo-workbook", type=Path, help="Planilha usada no piloto WooCommerce. Padrao: amostra de 5 ou produtos-novos.")
    parser.add_argument("--woo-limit", type=int, default=5, help="Quantidade de produtos no piloto WooCommerce.")
    parser.add_argument("--woo-status", choices=["draft", "publish", "private"], default="draft", help="Status dos produtos criados no piloto.")
    parser.add_argument("--woo-update-existing", action="store_true", help="Atualiza produto existente encontrado por SKU no piloto.")
    parser.add_argument("--woo-env-file", type=Path, help="Arquivo local com credenciais WooCommerce para o piloto.")
    parser.add_argument("--woo-site-url", help="URL base da loja WooCommerce; alternativa a WOOCOMMERCE_URL.")
    parser.add_argument("--woo-consumer-key", help="Consumer key WooCommerce; alternativa a WOOCOMMERCE_CONSUMER_KEY.")
    parser.add_argument("--woo-consumer-secret", help="Consumer secret WooCommerce; alternativa a WOOCOMMERCE_CONSUMER_SECRET.")
    parser.add_argument("--manifesto", type=Path, help="Arquivo JSON com resumo dos comandos, artefatos e exit codes.")
    parser.add_argument("--log-dir", type=Path, help="Pasta para gravar logs da execucao. Padrao: MSN/logs.")
    return parser.parse_args(argv)


def build_manifest(
    args: argparse.Namespace,
    conciliacao_folder: Path,
    candidate_root: Path,
    todos_workbook: Path,
    relatorio_workbook: Path,
    novos_workbook: Path,
    woo_pilot_workbook: Path,
    verification_output: Path,
    manifesto_path: Path,
    log_path: Path,
) -> dict[str, object]:
    return {
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "conciliacao_folder": str(conciliacao_folder),
        "candidate_root": str(candidate_root),
        "artifacts": {
            "todos_workbook": str(todos_workbook),
            "produtos_novos": str(novos_workbook),
            "relatorio": str(relatorio_workbook),
            "woo_pilot_workbook": str(woo_pilot_workbook),
            "verificacao": str(verification_output),
        },
        "steps": [],
        "args": manifest_args(args, conciliacao_folder, manifesto_path, log_path),
    }


def run_pipeline(args: argparse.Namespace, logger: object | None, log_path: Path) -> int:
    conciliacao_folder = args.conciliacao_folder.expanduser().resolve()
    candidate_root = DEFAULT_DESKTOP / args.desktop_folder_name
    candidate_root.mkdir(parents=True, exist_ok=True)
    todos_workbook = expected_conciliador_output(args, conciliacao_folder)
    relatorio_workbook = expected_relatorio_output(args, todos_workbook, conciliacao_folder)
    novos_workbook = expected_novos_output(args, todos_workbook, conciliacao_folder)
    woo_pilot_workbook = expected_woo_pilot_workbook(args, novos_workbook, conciliacao_folder)
    verification_output = expected_verification_output(conciliacao_folder)
    manifesto_path = args.manifesto or (conciliacao_folder / "manifesto-execucao.json")
    manifest = build_manifest(
        args,
        conciliacao_folder,
        candidate_root,
        todos_workbook,
        relatorio_workbook,
        novos_workbook,
        woo_pilot_workbook,
        verification_output,
        manifesto_path,
        log_path,
    )

    conciliador_command = build_conciliador_command(args, conciliacao_folder)
    code = run_script(conciliador_command, logger)
    record_step(manifest, "conciliador", conciliador_command, code)
    if code != 0:
        return finish_manifest(manifest, manifesto_path, code)
    manifest["summary"] = summarize_conciliacao_report(relatorio_workbook)

    if not args.skip_verify:
        verifier_command = build_verifier_command(args, conciliacao_folder, verification_output)
        code = run_script(verifier_command, logger)
        record_step(manifest, "verificador", verifier_command, code)
        if code != 0:
            return finish_manifest(manifest, manifesto_path, code)

    if args.woo_pilot:
        woo_command, manifest_command = build_woo_command(args, woo_pilot_workbook)
        code = run_script(woo_command, logger, display_command=manifest_command)
        record_step(manifest, "woo_pilot", manifest_command, code)
        if code != 0:
            return finish_manifest(manifest, manifesto_path, code)

    if args.skip_images:
        return finish_manifest(manifest, manifesto_path, 0)

    if not todos_workbook.exists():
        emit(logger, f"Erro: arquivo de conciliacao esperado nao encontrado: {todos_workbook}", error=True)
        return finish_manifest(manifest, manifesto_path, 1)

    buscador_command = build_buscador_command(args, conciliacao_folder, todos_workbook, candidate_root)
    code = run_script(buscador_command, logger)
    record_step(manifest, "buscador", buscador_command, code)
    if code != 0:
        return finish_manifest(manifest, manifesto_path, code)

    otimizador_command = build_otimizador_command(args, candidate_root, todos_workbook)
    code = run_script(otimizador_command, logger)
    record_step(manifest, "otimizador", otimizador_command, code)
    if code != 0:
        return finish_manifest(manifest, manifesto_path, code)

    return finish_manifest(manifest, manifesto_path, 0)


def main() -> int:
    args = parse_args()
    logger, log_path = setup_script_logging("run_all_scripts", SCRIPT_ROOT, args.log_dir)
    return run_pipeline(args, logger, log_path)


if __name__ == "__main__":
    raise SystemExit(main())
