from __future__ import annotations

import argparse
import csv
import html
import io
import json
import re
import subprocess
import sys
import time
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus, unquote, urljoin, urlparse

import requests


ROOT_DIR = Path(__file__).resolve().parent
PRODUCTS_DIR = ROOT_DIR / "products"
RAW_DIR = PRODUCTS_DIR / "_raw"
REPORTS_DIR = PRODUCTS_DIR / "_reports"
REPORT_FILE = REPORTS_DIR / "resultado-imagens.csv"
DEFAULT_DOWNLOADS_DIR = Path.home() / "Downloads"

DEFAULT_WORKBOOK_NAME = "Controle_de_estoque_Com_Filtro.xlsx"
DEFAULT_WORKSHEET_NAME = "Controle de Estoque"
DEFAULT_EXCEL_FILE = (
    Path.home()
    / "Desktop"
    / "Controle de Estoque"
    / DEFAULT_WORKBOOK_NAME
)

ICECAT_BASE_URL = "https://icecat.biz"
ICECAT_SEARCH_URL = f"{ICECAT_BASE_URL}/search"
PUBLIC_REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0 Safari/537.36 MSN-image-optimizer/1.0"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
}
FINAL_SIZE = (800, 800)
WEBP_QUALITY = 85
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
ZIP_EXTENSIONS = {".zip"}

BRAND_ALIASES = {
    "HP": "HP",
    "HEWLETT PACKARD": "HP",
    "HEWLETT-PACKARD": "HP",
    "SAMSUNG": "Samsung",
    "BROTHER": "Brother",
    "CANON": "Canon",
    "EPSON": "Epson",
    "LEXMARK": "Lexmark",
    "XEROX": "Xerox",
    "RICOH": "Ricoh",
    "KYOCERA": "Kyocera",
    "SHARP": "Sharp",
    "DELL": "Dell",
    "LENOVO": "Lenovo",
    "OKI": "OKI",
    "APC": "APC",
    "ZEBRA": "Zebra",
    "ELGIN": "Elgin",
}

SKU_SUFFIXES = {
    "AMR",
    "AZ",
    "CIA",
    "CIN",
    "MAG",
    "NV",
    "PRT",
    "PRETO",
    "CIANO",
    "AMARELO",
    "MAGENTA",
}

NOISE_CODES = {
    "TON",
    "TONER",
    "CART",
    "CARTUCHO",
    "CX",
    "NOVA",
    "ORIGINAL",
    "COMPATIVEL",
}


@dataclass(frozen=True)
class Product:
    sku: str
    name: str
    row_number: int | None = None
    brand: str | None = None
    product_code: str | None = None
    gtin: str | None = None
    source_dir: Path | None = None


@dataclass
class ProductResult:
    sku: str
    name: str
    status: str
    icecat_attempt: str = ""
    search_url: str = ""
    page_url: str = ""
    downloaded: int = 0
    processed: int = 0
    failed_images: int = 0
    message: str = ""
    raw_dir: str = ""
    output_dir: str = ""


def main() -> int:
    args = parse_args()
    PRODUCTS_DIR.mkdir(exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    if args.icecat_public:
        RAW_DIR.mkdir(parents=True, exist_ok=True)

    try:
        products = load_products(args)
    except Exception as exc:
        print(f"Erro ao carregar produtos: {exc}", file=sys.stderr)
        return 2

    if not products:
        if should_use_download_folders(args):
            print(
                "Nenhum arquivo ou pasta em Downloads com SKU correspondente na planilha. "
                f"Use nomes como: {args.downloads_dir.expanduser().resolve()}\\TON-HP-CE313AB-MAG "
                "ou TON-HP-CE313AB-MAG.zip"
            )
        else:
            print("Nenhum produto encontrado para processar.")
        write_report([])
        return 0

    if args.dry_run:
        return run_dry_run(products, args)

    client = PublicIcecatClient(args.search_delay) if args.icecat_public else None
    results: list[ProductResult] = []
    products_to_mark: list[Product] = []

    for index, product in enumerate(products, start=1):
        print(f"[{index}/{len(products)}] Produto {product.sku}: {product.name}")
        if args.input:
            result = process_local_product(product, args)
        elif args.manual_downloads:
            result = prepare_manual_download_product(product, args)
        elif args.icecat_public:
            assert client is not None
            result = process_public_icecat_product(product, client, args)
        else:
            result = process_downloads_product(product, args)
        results.append(result)
        print(
            f"  status={result.status} baixadas={result.downloaded} "
            f"processadas={result.processed} falhas={result.failed_images}"
        )
        if result.message:
            print(f"  detalhe={result.message}")
        if should_mark_excel_product(args, result):
            products_to_mark.append(product)

    if products_to_mark:
        try:
            marked = mark_excel_products_ok(products_to_mark, args)
            if marked:
                print(f"Planilha marcada: {marked} linha(s) com OK na coluna {args.mark_column}.")
        except Exception as exc:
            print(f"Aviso: nao foi possivel marcar OK na planilha: {exc}", file=sys.stderr)

    write_report(results)
    print(f"Relatorio gerado em: {REPORT_FILE}")

    failures = [item for item in results if item.status not in {"ok", "skipped_existing"}]
    if failures:
        print(f"Concluido com {len(failures)} produto(s) pendente(s).")
    else:
        print("Processamento concluido com sucesso.")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Importa imagens manuais de Downloads/[SKU] a partir do Excel aberto "
            "e otimiza para WooCommerce em WebP 800x800."
        )
    )
    parser.add_argument("--dry-run", action="store_true", help="Lista produtos e pastas sem baixar/processar.")
    parser.add_argument("--limit", type=int, help="Processa apenas os primeiros N produtos.")
    parser.add_argument("--sku", action="append", help="Processa somente um SKU especifico. Pode repetir.")
    parser.add_argument("--input", type=Path, help="Processa ZIPs/imagens locais sem buscar no Icecat.")
    parser.add_argument(
        "--downloads-dir",
        type=Path,
        default=DEFAULT_DOWNLOADS_DIR,
        help="Pasta base onde ficam as subpastas [SKU]. Padrao: ~/Downloads.",
    )
    parser.add_argument(
        "--manual-downloads",
        action="store_true",
        help="Cria pastas Downloads/[SKU] e abre/lista buscas do Icecat para download manual.",
    )
    parser.add_argument(
        "--icecat-public",
        action="store_true",
        help="Usa a busca publica do Icecat para tentar baixar imagens automaticamente.",
    )
    parser.add_argument(
        "--open-searches",
        action="store_true",
        help="Abre as buscas publicas do Icecat no navegador padrao.",
    )
    parser.add_argument(
        "--search-delay",
        type=float,
        default=4.0,
        help="Espera entre requisicoes publicas ao Icecat, em segundos.",
    )
    parser.add_argument("--max-images", type=int, help="Limita a quantidade de imagens baixadas por produto.")
    parser.add_argument("--skip-rembg", action="store_true", help="Nao remove fundo; apenas redimensiona e centraliza.")
    parser.add_argument(
        "--white-background",
        action="store_true",
        help="Usa fundo branco 800x800. Por padrao, o WebP final fica com fundo transparente.",
    )
    parser.add_argument("--overwrite", action="store_true", help="Recria imagens finais existentes.")
    parser.add_argument(
        "--mark-column",
        default="Imagem",
        help="Coluna da planilha que recebe OK quando o SKU em Downloads for tratado com sucesso.",
    )
    parser.add_argument(
        "--no-mark-excel",
        action="store_true",
        help="Nao escreve OK na planilha depois de tratar imagens.",
    )
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK_NAME, help="Nome do workbook aberto no Excel.")
    parser.add_argument("--sheet", default=DEFAULT_WORKSHEET_NAME, help="Nome da aba da planilha.")
    return parser.parse_args()


def load_products(args: argparse.Namespace) -> list[Product]:
    if args.input:
        products = products_from_local_input(args.input, args.sku)
    else:
        rows = read_visible_excel_rows(args.workbook, args.sheet)
        products = products_from_rows(rows)

    if args.sku:
        wanted = {normalize_sku(sku) for sku in args.sku}
        products = [product for product in products if normalize_sku(product.sku) in wanted]

    if should_use_download_folders(args):
        products = products_from_download_folders(args.downloads_dir, products)

    if args.limit is not None:
        products = products[: max(0, args.limit)]

    return products


def products_from_local_input(input_path: Path, sku_filters: list[str] | None) -> list[Product]:
    base = input_path.resolve()
    if not base.exists():
        raise FileNotFoundError(f"Pasta de entrada nao encontrada: {base}")

    wanted = {normalize_sku(sku) for sku in sku_filters or []}
    if base.name.lower() != "_raw" and any(is_supported_asset(path) for path in base.rglob("*") if path.is_file()):
        sku = sku_filters[0] if sku_filters else base.name
        return [Product(sku=safe_folder_name(sku), name=sku)]

    candidates: list[Path]
    if any(is_supported_asset(path) for path in base.iterdir() if path.is_file()):
        candidates = [base]
    else:
        candidates = [path for path in base.iterdir() if path.is_dir()]

    products: list[Product] = []
    for path in sorted(candidates):
        sku = path.name
        if wanted and normalize_sku(sku) not in wanted:
            continue
        products.append(Product(sku=safe_folder_name(sku), name=sku))
    return products


def products_from_rows(rows: list[dict[str, Any]]) -> list[Product]:
    products: list[Product] = []
    for row in rows:
        mapped = normalize_row_keys(row)
        sku = first_present(mapped, "sku", "codigo", "codigoproduto")
        name = first_present(mapped, "nome", "produto", "descricao")
        if not sku or not name:
            continue

        product = Product(
            sku=str(sku).strip(),
            name=str(name).strip(),
            row_number=to_int(mapped.get("__row")),
            brand=clean_optional(first_present(mapped, "marca", "brand")),
            product_code=clean_optional(
                first_present(
                    mapped,
                    "productcode",
                    "product_code",
                    "codigofabricante",
                    "codigoicecat",
                    "mpn",
                )
            ),
            gtin=clean_gtin(first_present(mapped, "gtin", "ean", "upc", "codigoean")),
        )
        products.append(product)
    return products


def products_from_download_folders(downloads_dir: Path, products: list[Product]) -> list[Product]:
    downloads_root = downloads_dir.expanduser().resolve()
    if not downloads_root.exists():
        return []

    products_by_sku = {normalize_sku(product.sku): product for product in products}
    matched: list[Product] = []
    seen: set[str] = set()
    for source_path in sorted(download_sources(downloads_root), key=download_source_sort_key):
        source_sku = source_path.name if source_path.is_dir() else source_path.stem
        normalized_sku = normalize_sku(source_sku)
        if normalized_sku in seen:
            continue
        product = products_by_sku.get(normalized_sku)
        if product is None:
            continue
        seen.add(normalized_sku)
        matched.append(
            Product(
                sku=product.sku,
                name=product.name,
                row_number=product.row_number,
                brand=product.brand,
                product_code=product.product_code,
                gtin=product.gtin,
                source_dir=source_path,
            )
        )
    return matched


def download_sources(downloads_root: Path) -> list[Path]:
    sources: list[Path] = []
    for path in downloads_root.iterdir():
        if path.is_dir():
            sources.append(path)
        elif path.is_file() and is_supported_asset(path):
            sources.append(path)
    return sources


def download_source_sort_key(path: Path) -> tuple[str, int, str]:
    source_sku = path.name if path.is_dir() else path.stem
    return normalize_sku(source_sku), 0 if path.is_dir() else 1, path.name.lower()


def read_visible_excel_rows(workbook_name: str, sheet_name: str) -> list[dict[str, Any]]:
    errors: list[str] = []
    for reader in (read_excel_rows_pywin32, read_excel_rows_powershell):
        try:
            rows = reader(workbook_name, sheet_name)
            if rows:
                return rows
        except Exception as exc:
            errors.append(f"{reader.__name__}: {exc}")

    try:
        rows = read_excel_rows_openpyxl(workbook_name, sheet_name)
        if rows:
            print(
                "Aviso: lendo a planilha salva no disco; filtros abertos no Excel "
                "podem nao ser refletidos sem pywin32/COM.",
                file=sys.stderr,
            )
            return rows
    except Exception as exc:
        errors.append(f"read_excel_rows_openpyxl: {exc}")

    raise RuntimeError("Nao foi possivel ler o Excel. " + " | ".join(errors))


def read_excel_rows_pywin32(workbook_name: str, sheet_name: str) -> list[dict[str, Any]]:
    try:
        import win32com.client  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("pywin32 nao instalado") from exc

    excel = win32com.client.GetActiveObject("Excel.Application")
    workbook = None
    for candidate in excel.Workbooks:
        if Path(candidate.FullName).name.lower() == workbook_name.lower() or candidate.Name.lower() == workbook_name.lower():
            workbook = candidate
            break
    if workbook is None:
        raise RuntimeError(f"workbook aberto nao encontrado: {workbook_name}")

    worksheet = workbook.Worksheets(sheet_name)
    used = worksheet.UsedRange
    row_count = int(used.Rows.Count)
    col_count = int(used.Columns.Count)
    headers = [str(worksheet.Cells(1, col).Text or "").strip() for col in range(1, col_count + 1)]

    rows: list[dict[str, Any]] = []
    for row_number in range(2, row_count + 1):
        if bool(worksheet.Rows(row_number).Hidden):
            continue
        item = {"__row": row_number}
        for col_number, header in enumerate(headers, start=1):
            if not header:
                continue
            item[header] = worksheet.Cells(row_number, col_number).Text
        rows.append(item)
    return rows


def read_excel_rows_powershell(workbook_name: str, sheet_name: str) -> list[dict[str, Any]]:
    ps_script = f"""
$ErrorActionPreference = 'Stop'
$workbookName = {json.dumps(workbook_name)}
$sheetName = {json.dumps(sheet_name)}
$excel = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application')
$target = $null
foreach ($wb in $excel.Workbooks) {{
  if ([System.IO.Path]::GetFileName($wb.FullName).ToLowerInvariant() -eq $workbookName.ToLowerInvariant() -or $wb.Name.ToLowerInvariant() -eq $workbookName.ToLowerInvariant()) {{
    $target = $wb
    break
  }}
}}
if (-not $target) {{ throw "workbook aberto nao encontrado: $workbookName" }}
$ws = $target.Worksheets.Item($sheetName)
$used = $ws.UsedRange
$rowCount = [int]$used.Rows.Count
$colCount = [int]$used.Columns.Count
$headers = @()
for ($c = 1; $c -le $colCount; $c++) {{
  $headers += [string]$ws.Cells.Item(1, $c).Text
}}
$rows = @()
for ($r = 2; $r -le $rowCount; $r++) {{
  if (-not $ws.Rows.Item($r).Hidden) {{
    $item = [ordered]@{{ "__row" = $r }}
    for ($c = 1; $c -le $colCount; $c++) {{
      $header = [string]$headers[$c - 1]
      if ($header.Trim().Length -gt 0) {{
        $item[$header] = [string]$ws.Cells.Item($r, $c).Text
      }}
    }}
    $rows += [pscustomobject]$item
  }}
}}
$rows | ConvertTo-Json -Depth 4 -Compress
"""
    completed = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        detail = completed.stderr.strip() or completed.stdout.strip()
        raise RuntimeError(detail)
    output = completed.stdout.strip()
    if not output:
        return []
    data = json.loads(output)
    if isinstance(data, dict):
        return [data]
    return data


def read_excel_rows_openpyxl(workbook_name: str, sheet_name: str) -> list[dict[str, Any]]:
    import openpyxl

    workbook_path = DEFAULT_EXCEL_FILE
    if workbook_path.name.lower() != workbook_name.lower():
        workbook_path = DEFAULT_EXCEL_FILE.with_name(workbook_name)
    if not workbook_path.exists():
        raise FileNotFoundError(f"arquivo nao encontrado: {workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    worksheet = workbook[sheet_name]
    rows_iter = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows_iter)]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows_iter, start=2):
        dimension = worksheet.row_dimensions[row_number]
        if bool(dimension.hidden):
            continue
        item = {"__row": row_number}
        for header, value in zip(headers, values):
            if header:
                item[header] = value
        rows.append(item)
    workbook.close()
    return rows


def mark_excel_products_ok(products: list[Product], args: argparse.Namespace) -> int:
    rows = [product.row_number for product in products if product.row_number is not None]
    if not rows:
        return 0

    errors: list[str] = []
    for writer in (mark_excel_products_ok_pywin32, mark_excel_products_ok_powershell):
        try:
            return writer(args.workbook, args.sheet, rows, args.mark_column)
        except Exception as exc:
            errors.append(f"{writer.__name__}: {exc}")

    try:
        return mark_excel_products_ok_openpyxl(args.workbook, args.sheet, rows, args.mark_column)
    except Exception as exc:
        errors.append(f"mark_excel_products_ok_openpyxl: {exc}")

    raise RuntimeError(" | ".join(errors))


def mark_excel_products_ok_pywin32(
    workbook_name: str,
    sheet_name: str,
    row_numbers: list[int],
    column_name: str,
) -> int:
    try:
        import win32com.client  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("pywin32 nao instalado") from exc

    excel = win32com.client.GetActiveObject("Excel.Application")
    workbook = None
    for candidate in excel.Workbooks:
        if Path(candidate.FullName).name.lower() == workbook_name.lower() or candidate.Name.lower() == workbook_name.lower():
            workbook = candidate
            break
    if workbook is None:
        raise RuntimeError(f"workbook aberto nao encontrado: {workbook_name}")

    worksheet = workbook.Worksheets(sheet_name)
    status_col = ensure_excel_status_column_pywin32(worksheet, column_name)
    for row_number in row_numbers:
        worksheet.Cells(int(row_number), status_col).Value = "OK"
    return len(row_numbers)


def ensure_excel_status_column_pywin32(worksheet: Any, column_name: str) -> int:
    used = worksheet.UsedRange
    col_count = int(used.Columns.Count)
    target = normalize_header(column_name)
    for col in range(1, col_count + 1):
        header = str(worksheet.Cells(1, col).Text or "").strip()
        if normalize_header(header) == target:
            return col

    status_col = col_count + 1
    worksheet.Cells(1, status_col).Value = column_name
    return status_col


def mark_excel_products_ok_powershell(
    workbook_name: str,
    sheet_name: str,
    row_numbers: list[int],
    column_name: str,
) -> int:
    ps_script = f"""
$ErrorActionPreference = 'Stop'
$workbookName = {json.dumps(workbook_name)}
$sheetName = {json.dumps(sheet_name)}
$columnName = {json.dumps(column_name)}
$rowsJson = {json.dumps(json.dumps(row_numbers))}
$rows = @($rowsJson | ConvertFrom-Json)
$excel = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application')
$target = $null
foreach ($wb in $excel.Workbooks) {{
  if ([System.IO.Path]::GetFileName($wb.FullName).ToLowerInvariant() -eq $workbookName.ToLowerInvariant() -or $wb.Name.ToLowerInvariant() -eq $workbookName.ToLowerInvariant()) {{
    $target = $wb
    break
  }}
}}
if (-not $target) {{ throw "workbook aberto nao encontrado: $workbookName" }}
$ws = $target.Worksheets.Item($sheetName)
$used = $ws.UsedRange
$colCount = [int]$used.Columns.Count
$statusCol = 0
for ($c = 1; $c -le $colCount; $c++) {{
  $header = [string]$ws.Cells.Item(1, $c).Text
  if ($header.Trim().ToLowerInvariant() -eq $columnName.Trim().ToLowerInvariant()) {{
    $statusCol = $c
    break
  }}
}}
if ($statusCol -eq 0) {{
  $statusCol = $colCount + 1
  $ws.Cells.Item(1, $statusCol).Value2 = $columnName
}}
foreach ($row in $rows) {{
  $ws.Cells.Item([int]$row, $statusCol).Value2 = 'OK'
}}
$rows.Count
"""
    completed = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        detail = completed.stderr.strip() or completed.stdout.strip()
        raise RuntimeError(detail)
    output = completed.stdout.strip()
    return int(output) if output else 0


def mark_excel_products_ok_openpyxl(
    workbook_name: str,
    sheet_name: str,
    row_numbers: list[int],
    column_name: str,
) -> int:
    import openpyxl

    workbook_path = DEFAULT_EXCEL_FILE
    if workbook_path.name.lower() != workbook_name.lower():
        workbook_path = DEFAULT_EXCEL_FILE.with_name(workbook_name)
    if not workbook_path.exists():
        raise FileNotFoundError(f"arquivo nao encontrado: {workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[sheet_name]
    target = normalize_header(column_name)
    status_col = 0
    for col in range(1, worksheet.max_column + 1):
        header = str(worksheet.cell(row=1, column=col).value or "").strip()
        if normalize_header(header) == target:
            status_col = col
            break
    if status_col == 0:
        status_col = worksheet.max_column + 1
        worksheet.cell(row=1, column=status_col, value=column_name)

    for row_number in row_numbers:
        worksheet.cell(row=int(row_number), column=status_col, value="OK")
    workbook.save(workbook_path)
    workbook.close()
    return len(row_numbers)


class PublicIcecatClient:
    def __init__(self, search_delay: float) -> None:
        self.search_delay = max(0.0, search_delay)
        self.session = requests.Session()
        self.last_request_at = 0.0

    def find_images(self, product: Product) -> tuple[str, str, str, list[str]]:
        attempts = build_public_search_attempts(product)
        last_error = ""
        for attempt in attempts:
            search_url = public_search_url(attempt)
            try:
                search_html = self.get_text(search_url)
            except RateLimitedError:
                return attempt, search_url, "", []
            except Exception as exc:
                last_error = f"erro na busca publica: {exc}"
                continue

            page_urls = parse_product_page_links(search_html, search_url)
            page_urls = sort_candidate_pages(page_urls, product)
            if not page_urls:
                last_error = "busca publica sem paginas de produto"
                continue

            for page_url in page_urls[:4]:
                try:
                    product_html = self.get_text(page_url)
                except RateLimitedError:
                    return attempt, search_url, page_url, []
                except Exception as exc:
                    last_error = f"erro ao abrir pagina publica: {exc}"
                    continue

                image_urls = extract_public_image_urls(product_html)
                if image_urls:
                    return attempt, search_url, page_url, image_urls
                last_error = "pagina publica sem imagens detectadas"

        return attempts[0] if attempts else product.name, public_search_url(attempts[0] if attempts else product.name), "", []

    def download(self, url: str, destination: Path) -> None:
        destination.parent.mkdir(parents=True, exist_ok=True)
        response = self.get_response(url, stream=True, timeout=60)
        with destination.open("wb") as output:
            for chunk in response.iter_content(chunk_size=1024 * 256):
                if chunk:
                    output.write(chunk)

    def get_text(self, url: str) -> str:
        return self.get_response(url, timeout=30).text

    def get_response(self, url: str, **kwargs: Any) -> requests.Response:
        self.wait_before_request()
        response = self.session.get(url, headers=PUBLIC_REQUEST_HEADERS, **kwargs)
        self.last_request_at = time.monotonic()
        if response.status_code == 429:
            raise RateLimitedError("Icecat retornou 429 Too Many Requests")
        response.raise_for_status()
        return response

    def wait_before_request(self) -> None:
        elapsed = time.monotonic() - self.last_request_at
        wait_for = self.search_delay - elapsed
        if wait_for > 0:
            time.sleep(wait_for)


class RateLimitedError(RuntimeError):
    pass


def process_public_icecat_product(product: Product, client: PublicIcecatClient, args: argparse.Namespace) -> ProductResult:
    raw_product_dir = RAW_DIR / safe_folder_name(product.sku)
    downloaded_dir = raw_product_dir / "downloaded"
    output_dir = PRODUCTS_DIR / safe_folder_name(product.sku)
    result = ProductResult(
        sku=product.sku,
        name=product.name,
        status="pending",
        raw_dir=str(raw_product_dir),
        output_dir=str(output_dir),
    )

    attempt, search_url, page_url, image_urls = client.find_images(product)
    result.icecat_attempt = attempt
    result.search_url = search_url
    result.page_url = page_url

    if not image_urls:
        result.status = "not_found"
        result.message = (
            "sem imagens publicas detectadas; salve ZIPs/imagens em "
            f"\"{args.downloads_dir.expanduser().resolve() / safe_folder_name(product.sku)}\" "
            "e rode o script novamente"
        )
        return result

    if args.max_images is not None:
        image_urls = image_urls[: max(0, args.max_images)]

    last_error = ""
    for image_index, url in enumerate(image_urls, start=1):
        extension = extension_from_url(url) or ".jpg"
        destination = downloaded_dir / f"icecat-{image_index:02d}{extension}"
        if destination.exists() and not args.overwrite:
            result.downloaded += 1
            continue
        try:
            client.download(url, destination)
            result.downloaded += 1
        except Exception as exc:
            result.failed_images += 1
            last_error = f"falha ao baixar imagem {image_index}: {exc}"

    if result.downloaded == 0:
        result.status = "no_images"
        result.message = last_error or "nenhuma URL de imagem baixavel encontrada"
        return result

    processed, failed, message = process_image_folder(raw_product_dir, output_dir, product.sku, args)
    result.processed = processed
    result.failed_images += failed
    result.message = message or last_error
    result.status = status_from_counts(processed, result.failed_images)
    return result


def process_downloads_product(product: Product, args: argparse.Namespace) -> ProductResult:
    downloads_root = args.downloads_dir.expanduser().resolve()
    source_dir = product.source_dir or downloads_root / safe_folder_name(product.sku)
    output_dir = PRODUCTS_DIR / safe_folder_name(product.sku)
    if not source_dir.exists():
        return ProductResult(
            sku=product.sku,
            name=product.name,
            status="manual_missing",
            downloaded=0,
            processed=0,
            failed_images=0,
            message=(
                "arquivo ou pasta de download nao encontrado; salve ZIPs/imagens em "
                f"\"{source_dir}\" ou \"{source_dir}.zip\" e rode o script novamente"
            ),
            raw_dir=str(source_dir),
            output_dir=str(output_dir),
        )
    if not source_dir.is_dir() and not source_dir.is_file():
        return ProductResult(
            sku=product.sku,
            name=product.name,
            status="error",
            downloaded=0,
            processed=0,
            failed_images=1,
            message=f"origem esperada como arquivo ou pasta, mas nao e suportada: {source_dir}",
            raw_dir=str(source_dir),
            output_dir=str(output_dir),
        )

    processed, failed, message = process_image_source(source_dir, output_dir, product.sku, args)
    return ProductResult(
        sku=product.sku,
        name=product.name,
        status=status_from_counts(processed, failed),
        downloaded=0,
        processed=processed,
        failed_images=failed,
        message=message,
        raw_dir=str(source_dir),
        output_dir=str(output_dir),
    )


def process_local_product(product: Product, args: argparse.Namespace) -> ProductResult:
    assert args.input is not None
    input_root = args.input.resolve()
    product_input = input_root
    if not any(is_supported_asset(path) for path in input_root.iterdir() if path.is_file()):
        child = input_root / safe_folder_name(product.sku)
        if child.exists():
            product_input = child

    output_dir = PRODUCTS_DIR / safe_folder_name(product.sku)
    processed, failed, message = process_image_folder(product_input, output_dir, product.sku, args)
    return ProductResult(
        sku=product.sku,
        name=product.name,
        status=status_from_counts(processed, failed),
        downloaded=0,
        processed=processed,
        failed_images=failed,
        message=message,
        raw_dir=str(product_input),
        output_dir=str(output_dir),
    )


def prepare_manual_download_product(product: Product, args: argparse.Namespace) -> ProductResult:
    manual_product_dir = args.downloads_dir.expanduser().resolve() / safe_folder_name(product.sku)
    manual_product_dir.mkdir(parents=True, exist_ok=True)
    search_terms = build_public_search_attempts(product)
    search_url = public_search_url(search_terms[0] if search_terms else product.name)

    if args.open_searches:
        import webbrowser

        webbrowser.open(search_url)

    return ProductResult(
        sku=product.sku,
        name=product.name,
        status="manual_pending",
        icecat_attempt=search_terms[0] if search_terms else product.name,
        search_url=search_url,
        downloaded=0,
        processed=0,
        failed_images=0,
        message=(
            "pasta criada para download manual; salve ZIPs/imagens em "
            f"\"{manual_product_dir}\" e rode python otimizador_imagens.py "
            f"--sku \"{product.sku}\""
        ),
        raw_dir=str(manual_product_dir),
        output_dir=str(PRODUCTS_DIR / safe_folder_name(product.sku)),
    )


def process_image_folder(
    source_dir: Path,
    output_dir: Path,
    sku: str,
    args: argparse.Namespace,
) -> tuple[int, int, str]:
    if not source_dir.exists():
        return 0, 0, f"pasta de origem nao encontrada: {source_dir}"

    extracted_dir = source_dir / "extracted"
    zip_messages = extract_zip_files(source_dir, extracted_dir, overwrite=args.overwrite)
    image_paths = collect_image_files(source_dir, output_dir=output_dir, sku=sku)
    if not image_paths:
        return 0, 0, "; ".join(zip_messages + ["nenhuma imagem encontrada"])

    return process_image_paths(image_paths, output_dir, sku, args, zip_messages)


def process_image_source(
    source_path: Path,
    output_dir: Path,
    sku: str,
    args: argparse.Namespace,
) -> tuple[int, int, str]:
    if source_path.is_dir():
        return process_image_folder(source_path, output_dir, sku, args)

    suffix = source_path.suffix.lower()
    if suffix in ZIP_EXTENSIONS:
        source_dir = source_path.parent / safe_folder_name(source_path.stem)
        extracted_dir = source_dir / "extracted"
        zip_messages = extract_single_zip_file(source_path, extracted_dir, overwrite=args.overwrite)
        image_paths = collect_image_files(source_dir, output_dir=output_dir, sku=sku)
        if not image_paths:
            return 0, 0, "; ".join(zip_messages + ["nenhuma imagem encontrada"])
        return process_image_paths(image_paths, output_dir, sku, args, zip_messages)

    if suffix in IMAGE_EXTENSIONS:
        return process_image_paths([source_path], output_dir, sku, args, [])

    return 0, 1, f"arquivo nao suportado: {source_path.name}"


def process_image_paths(
    image_paths: list[Path],
    output_dir: Path,
    sku: str,
    args: argparse.Namespace,
    messages: list[str] | None = None,
) -> tuple[int, int, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    processed = 0
    failed = 0
    output_messages = list(messages or [])

    for index, image_path in enumerate(image_paths, start=1):
        destination = output_dir / f"{safe_folder_name(sku)}-{index:02d}.webp"
        if destination.exists() and not args.overwrite:
            processed += 1
            continue
        try:
            optimize_image(
                image_path,
                destination,
                skip_rembg=args.skip_rembg,
                white_background=args.white_background,
            )
            processed += 1
        except Exception as exc:
            failed += 1
            output_messages.append(f"{image_path.name}: {exc}")

    return processed, failed, "; ".join(output_messages)


def optimize_image(image_path: Path, destination: Path, skip_rembg: bool, white_background: bool) -> None:
    from PIL import Image, ImageOps

    if skip_rembg:
        with Image.open(image_path) as original:
            image = ImageOps.exif_transpose(original).convert("RGBA")
    else:
        from rembg import remove

        input_bytes = image_path.read_bytes()
        output_bytes = remove(input_bytes)
        image = Image.open(io.BytesIO(output_bytes)).convert("RGBA")

    image.thumbnail(FINAL_SIZE, Image.Resampling.LANCZOS)
    position = (
        (FINAL_SIZE[0] - image.width) // 2,
        (FINAL_SIZE[1] - image.height) // 2,
    )
    if white_background:
        canvas = Image.new("RGB", FINAL_SIZE, (255, 255, 255))
        canvas.paste(image, position, image)
    else:
        canvas = Image.new("RGBA", FINAL_SIZE, (0, 0, 0, 0))
        canvas.alpha_composite(image, position)
    destination.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(destination, "webp", quality=WEBP_QUALITY, method=6)


def extract_zip_files(source_dir: Path, extracted_dir: Path, overwrite: bool = False) -> list[str]:
    messages: list[str] = []
    extracted_root = extracted_dir.resolve()
    for zip_path in sorted(source_dir.rglob("*")):
        if not zip_path.is_file() or zip_path.suffix.lower() not in ZIP_EXTENSIONS:
            continue
        try:
            zip_path.resolve().relative_to(extracted_root)
            continue
        except ValueError:
            pass

        messages.extend(extract_single_zip_file(zip_path, extracted_dir, overwrite=overwrite))
    return messages


def extract_single_zip_file(zip_path: Path, extracted_dir: Path, overwrite: bool = False) -> list[str]:
    messages: list[str] = []
    target_dir = extracted_dir / safe_folder_name(zip_path.stem)
    if target_dir.exists() and not overwrite and contains_image_files(target_dir):
        messages.append(f"{zip_path.name}: extracao pulada; pasta extracted ja contem imagens")
        return messages
    try:
        safe_extract_zip(zip_path, target_dir)
    except Exception as exc:
        messages.append(f"erro ao extrair {zip_path.name}: {exc}")
    return messages


def safe_extract_zip(zip_path: Path, target_dir: Path) -> None:
    target_dir.mkdir(parents=True, exist_ok=True)
    target_root = target_dir.resolve()
    with zipfile.ZipFile(zip_path) as archive:
        for member in archive.infolist():
            destination = (target_dir / member.filename).resolve()
            try:
                destination.relative_to(target_root)
            except ValueError as exc:
                raise RuntimeError(f"entrada insegura no ZIP: {member.filename}")
        archive.extractall(target_dir)


def collect_image_files(source_dir: Path, output_dir: Path | None = None, sku: str | None = None) -> list[Path]:
    images: list[Path] = []
    safe_sku = safe_folder_name(sku or "").lower()
    output_root = output_dir.resolve() if output_dir else None
    for path in sorted(source_dir.rglob("*")):
        if not path.is_file():
            continue
        if output_root and safe_sku:
            try:
                path.parent.resolve().relative_to(output_root)
                if path.suffix.lower() == ".webp" and path.stem.lower().startswith(f"{safe_sku}-"):
                    continue
            except ValueError:
                pass
        if path.suffix.lower() in IMAGE_EXTENSIONS:
            images.append(path)
    return images


def contains_image_files(source_dir: Path) -> bool:
    return any(path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS for path in source_dir.rglob("*"))


def build_public_search_attempts(product: Product) -> list[str]:
    attempts: list[str] = []

    def add(value: str | None) -> None:
        if not value:
            return
        normalized = re.sub(r"\s+", " ", value).strip()
        if normalized and normalized not in attempts:
            attempts.append(normalized)

    brand = product.brand or infer_brand(product)
    product_code = clean_product_code(product.product_code) or infer_product_code(product)
    if product.gtin:
        add(product.gtin)
    if brand and product_code:
        add(f"{brand} {product_code}")
    if product_code:
        add(product_code)
    add(product.name)
    add(product.sku)
    return attempts


def public_search_url(search_term: str) -> str:
    return f"{ICECAT_SEARCH_URL}?keyword={quote_plus(search_term)}"


def parse_product_page_links(page_html: str, base_url: str) -> list[str]:
    links: list[str] = []
    for match in re.finditer(r"""href\s*=\s*["']([^"']+)["']""", page_html, flags=re.IGNORECASE):
        href = html.unescape(match.group(1))
        href = href.replace("\\u002F", "/")
        url = urljoin(base_url, href)
        parsed = urlparse(url)
        if "icecat.biz" not in parsed.netloc.lower():
            continue
        decoded_path = unquote(parsed.path).lower()
        if "/p/" not in decoded_path or not decoded_path.endswith(".html"):
            continue
        clean_url = url.split("#", 1)[0]
        if clean_url not in links:
            links.append(clean_url)
    return links


def sort_candidate_pages(urls: list[str], product: Product) -> list[str]:
    return sorted(urls, key=lambda url: score_product_page_url(url, product), reverse=True)


def score_product_page_url(url: str, product: Product) -> int:
    decoded = unquote(url).upper()
    score = 0
    brand = product.brand or infer_brand(product)
    code = clean_product_code(product.product_code) or infer_product_code(product)
    if code and code.upper() in decoded:
        score += 8
    if normalize_sku(product.sku) in decoded:
        score += 5
    if brand and brand.upper() in decoded:
        score += 3
    for token in re.findall(r"[A-Z0-9]{4,}", normalize_code_text(product.name)):
        if token in decoded:
            score += 1
    return score


def extract_public_image_urls(page_html: str) -> list[str]:
    matches = re.findall(
        r"""https://images\.icecat\.biz/[^"'<>\\\s)]+""",
        page_html,
        flags=re.IGNORECASE,
    )
    ranked: dict[str, tuple[int, str]] = {}
    for raw_url in matches:
        url = html.unescape(raw_url)
        url = url.replace("\\u002F", "/")
        url = url.rstrip(".,;)")
        parsed = urlparse(url)
        suffix = Path(parsed.path).suffix.lower()
        if suffix not in IMAGE_EXTENSIONS.union(ZIP_EXTENSIONS):
            continue
        if "/img/gallery" not in parsed.path and "/img/norm/" not in parsed.path:
            continue

        key = public_image_key(parsed.path)
        priority = public_image_priority(parsed.path)
        current = ranked.get(key)
        if current is None or priority > current[0]:
            ranked[key] = (priority, url)

    return [url for _, url in sorted(ranked.values(), key=lambda item: item[1])]


def public_image_key(path: str) -> str:
    name = Path(path).name
    stem = Path(name).stem
    stem = re.sub(r"_(?:low|medium|thumb|large)$", "", stem, flags=re.IGNORECASE)
    return stem.lower()


def public_image_priority(path: str) -> int:
    lower_path = path.lower()
    if "/gallery_raw/" in lower_path:
        return 50
    if "/gallery/" in lower_path:
        return 40
    if "/gallery_mediums/" in lower_path:
        return 30
    if "/gallery_lows/" in lower_path:
        return 20
    if "/gallery_thumbs/" in lower_path:
        return 10
    return 1


def infer_brand(product: Product) -> str | None:
    text = f"{product.name} {product.sku}".upper()
    text = re.sub(r"[_/-]+", " ", text)
    for alias, canonical in BRAND_ALIASES.items():
        if re.search(rf"(^|\s){re.escape(alias)}($|\s)", text):
            return canonical
    return None


def infer_product_code(product: Product) -> str | None:
    for source in (product.name, product.sku):
        candidates = find_code_candidates(source)
        if candidates:
            return candidates[0]
    return None


def find_code_candidates(text: str) -> list[str]:
    normalized = normalize_code_text(text)
    patterns = [
        r"\b[A-Z]{1,5}-[A-Z]?\d[A-Z0-9]{2,}(?:-[A-Z0-9]+)?\b",
        r"\b[A-Z]{1,5}\d[A-Z0-9]{2,}\b",
        r"\b\d{4,}[A-Z]{1,4}\b",
    ]
    candidates: list[str] = []
    for pattern in patterns:
        for match in re.findall(pattern, normalized):
            code = clean_product_code(match)
            if code and code not in candidates:
                candidates.append(code)

    if candidates:
        return candidates

    pieces = re.split(r"[-_\s/]+", normalized)
    for piece in pieces:
        code = clean_product_code(piece)
        if code and code not in candidates:
            candidates.append(code)
    return candidates


def clean_product_code(value: Any) -> str | None:
    if value is None:
        return None
    code = str(value).strip().upper()
    if not code:
        return None
    code = normalize_code_text(code)
    parts = [part for part in re.split(r"[-_/]+", code) if part]
    while len(parts) > 1 and any(has_digit(part) for part in parts[:-1]) and not has_digit(parts[-1]):
        parts.pop()
    code = "-".join(parts)
    code = code.strip("-_ /")
    if not has_digit(code):
        return None
    if code in NOISE_CODES or code in SKU_SUFFIXES:
        return None
    if len(code) < 4:
        return None
    return code


def normalize_code_text(value: str) -> str:
    text = strip_accents(value).upper()
    text = re.sub(r"[^A-Z0-9_/\-\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def has_digit(value: str) -> bool:
    return any(char.isdigit() for char in value)


def extension_from_url(url: str) -> str | None:
    path = urlparse(url).path
    suffix = Path(path).suffix.lower()
    if suffix in IMAGE_EXTENSIONS or suffix in ZIP_EXTENSIONS:
        return suffix
    return None


def is_supported_asset(path: Path) -> bool:
    return path.suffix.lower() in IMAGE_EXTENSIONS.union(ZIP_EXTENSIONS)


def status_from_counts(processed: int, failed: int) -> str:
    if processed > 0 and failed == 0:
        return "ok"
    if processed > 0 and failed > 0:
        return "partial"
    if failed > 0:
        return "error"
    return "no_images"


def should_use_download_folders(args: argparse.Namespace) -> bool:
    return not args.input and not args.manual_downloads and not args.icecat_public


def should_mark_excel_product(args: argparse.Namespace, result: ProductResult) -> bool:
    if args.no_mark_excel or not should_use_download_folders(args):
        return False
    return result.status in {"ok", "skipped_existing"}


def run_dry_run(products: list[Product], args: argparse.Namespace) -> int:
    print(f"Dry-run: {len(products)} produto(s) seriam processados.")
    rows: list[ProductResult] = []
    for product in products:
        attempts = build_public_search_attempts(product)
        attempt_text = " | ".join(attempts) or "sem tentativa"
        search_url = ""
        if args.manual_downloads or args.icecat_public:
            search_url = public_search_url(attempts[0] if attempts else product.name)
        if args.input:
            raw_dir = str(args.input.expanduser().resolve())
        elif product.source_dir:
            raw_dir = str(product.source_dir)
        else:
            raw_dir = str(args.downloads_dir.expanduser().resolve() / safe_folder_name(product.sku))
        output_dir = str(PRODUCTS_DIR / safe_folder_name(product.sku))
        print(f"- {product.sku} :: {product.name} :: {attempt_text}")
        print(f"  origem: {raw_dir}")
        print(f"  destino: {output_dir}")
        if search_url:
            print(f"  busca: {search_url}")
        rows.append(
            ProductResult(
                sku=product.sku,
                name=product.name,
                status="dry_run",
                icecat_attempt=attempt_text,
                search_url=search_url,
                raw_dir=raw_dir,
                output_dir=output_dir,
                message="simulacao; nada foi baixado ou processado",
            )
        )
    write_report(rows)
    print(f"Relatorio dry-run gerado em: {REPORT_FILE}")
    return 0


def write_report(results: list[ProductResult]) -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    fields = [
        "sku",
        "name",
        "status",
        "icecat_attempt",
        "search_url",
        "page_url",
        "downloaded",
        "processed",
        "failed_images",
        "message",
        "raw_dir",
        "output_dir",
    ]
    with REPORT_FILE.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for result in results:
            writer.writerow({field: getattr(result, field) for field in fields})


def normalize_row_keys(row: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for key, value in row.items():
        normalized = normalize_header(str(key))
        mapped[normalized] = value
    return mapped


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9_]", "", strip_accents(value).lower())


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def clean_optional(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def clean_gtin(value: Any) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if 8 <= len(digits) <= 14:
        return digits
    return None


def to_int(value: Any) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return None


def safe_folder_name(value: str) -> str:
    text = strip_accents(str(value)).strip()
    text = re.sub(r"[<>:\"/\\|?*\x00-\x1F]", "-", text)
    text = re.sub(r"\s+", " ", text)
    text = text.strip(" .")
    return text or "produto"


def normalize_sku(value: str) -> str:
    return safe_folder_name(value).upper()


if __name__ == "__main__":
    raise SystemExit(main())
