from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import mimetypes
import os
import shutil
import re
import time
import unicodedata
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote_plus, unquote, urlencode, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent
PRODUCTS_DIR = ROOT_DIR / "products"
LEGACY_CANDIDATES_DIR = PRODUCTS_DIR / "_candidates"
APPROVED_RAW_DIR = PRODUCTS_DIR / "_approved_raw"
REPORTS_DIR = PRODUCTS_DIR / "_reports"
CANDIDATES_CSV = REPORTS_DIR / "candidatas-imagens.csv"
APPROVAL_TEMPLATE_CSV = REPORTS_DIR / "candidatas-aprovacao.csv"
SKIPPED_PRODUCTS_CSV = REPORTS_DIR / "skus-pulados-imagens.csv"
REVIEW_HTML = REPORTS_DIR / "revisao-candidatas.html"

DEFAULT_SOURCE_ROOT = Path.home() / "Desktop" / "cópia de produtos" / "Produtos"
DEFAULT_WORKBOOK_NAME = "Controle_de_estoque_Com_Filtro.xlsx"
DEFAULT_WORKBOOK = DEFAULT_SOURCE_ROOT / DEFAULT_WORKBOOK_NAME
DEFAULT_SHEET = "Controle de Estoque"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0 Safari/537.36 MSN-image-candidate-review/1.0"
)

BRAND_ALIASES = {
    "HP": ["HP", "HEWLETT PACKARD", "HEWLETT-PACKARD"],
    "Samsung": ["SAMSUNG"],
    "Brother": ["BROTHER"],
    "Canon": ["CANON"],
    "Epson": ["EPSON"],
    "Lexmark": ["LEXMARK"],
    "Xerox": ["XEROX"],
    "Ricoh": ["RICOH"],
    "Kyocera": ["KYOCERA"],
    "Sharp": ["SHARP"],
    "Dell": ["DELL"],
    "Lenovo": ["LENOVO"],
    "OKI": ["OKI"],
    "APC": ["APC"],
    "Zebra": ["ZEBRA"],
    "Elgin": ["ELGIN"],
}

BRAND_DOMAINS = {
    "HP": ["hp.com", "support.hp.com", "www8-hp.com", "hpe.com"],
    "Samsung": ["samsung.com"],
    "Brother": ["brother.com", "brother-usa.com", "brother.com.br"],
    "Canon": ["canon.com", "canon.com.br"],
    "Epson": ["epson.com", "epson.com.br"],
    "Lexmark": ["lexmark.com"],
    "Xerox": ["xerox.com"],
    "Ricoh": ["ricoh.com"],
    "Kyocera": ["kyoceradocumentsolutions.com", "kyocera.com"],
    "Sharp": ["sharpusa.com", "global.sharp"],
    "Dell": ["dell.com"],
    "Lenovo": ["lenovo.com"],
    "OKI": ["oki.com"],
    "APC": ["apc.com", "se.com"],
    "Zebra": ["zebra.com"],
    "Elgin": ["elgin.com.br"],
}

KIND_SYNONYMS = {
    "toner": ["TONER", "TONER CARTRIDGE", "CARTUCHO DE TONER"],
    "cartucho": ["CARTUCHO", "INK CARTRIDGE", "PRINT CARTRIDGE"],
    "drum": ["DRUM", "DRUM UNIT", "UNIDADE DE TAMBOR", "TAMBOR", "CILINDRO"],
    "impressora": ["IMPRESSORA", "PRINTER", "MULTIFUNCIONAL", "MFP"],
    "fusor": ["FUSOR", "FUSER"],
    "residuo": ["RESIDUO", "RESIDUAL", "WASTE", "WASTE TONER"],
    "developer": ["DEVELOPER", "REVELADOR"],
    "kit_manutencao": ["KIT MANUTENCAO", "KIT DE MANUTENCAO", "MAINTENANCE KIT"],
    "rolete": ["ROLETE", "ROLETES", "ROLO", "ROLLER"],
    "transferencia": ["TRANSFERENCIA", "TRANSFER", "BELT", "CORREIA", "CINTA"],
    "grampo": ["GRAMPO", "GRAMPOS", "STAPLE"],
    "papel": ["PAPEL", "SULFITE", "FILICOAT"],
    "tinta": ["TINTA", "INK"],
    "cera": ["CERA", "COLORQUBE"],
}

KIND_LABELS = {
    "toner": "toner",
    "cartucho": "cartucho",
    "drum": "drum",
    "impressora": "impressora",
    "fusor": "fusor",
    "residuo": "waste toner",
    "developer": "developer",
    "kit_manutencao": "kit manutencao",
    "rolete": "rolete",
    "transferencia": "transfer belt",
    "grampo": "grampo",
    "papel": "papel",
    "tinta": "tinta",
    "cera": "cera",
}

COLOR_SYNONYMS = {
    "preto": ["PRETO", "BLACK", "NEGRO", "NERO", "NOIR", "SCHWARZ", "CZARNY"],
    "ciano": ["CIANO", "CYAN"],
    "magenta": ["MAGENTA"],
    "amarelo": ["AMARELO", "YELLOW", "AMARILLO", "JAUNE", "GELB"],
}

COLOR_SUFFIXES = {
    "PRT": "preto",
    "PRETO": "preto",
    "BK": "preto",
    "CIA": "ciano",
    "CIANO": "ciano",
    "CYAN": "ciano",
    "MAG": "magenta",
    "MAGENTA": "magenta",
    "AMR": "amarelo",
    "AMARELO": "amarelo",
    "YELLOW": "amarelo",
}

PACKAGE_SYNONYMS = {
    "caixa_azul_branca": [
        "CX AZUL E BRANCA",
        "CX AZUL BRANCA",
        "CAIXA AZUL E BRANCA",
        "CAIXA AZUL BRANCA",
        "AZUL E BRANCA",
        "AZUL BRANCA",
        "BLUE WHITE BOX",
        "BLUE AND WHITE BOX",
    ],
    "caixa_nova_preta": [
        "CX NOVA PRETA",
        "CAIXA NOVA PRETA",
        "CX PRETA",
        "CAIXA PRETA",
        "NOVA PRETA",
        "NEW BLACK BOX",
        "BLACK BOX",
    ],
    "caixa_branca": [
        "CX BRANCA",
        "CAIXA BRANCA",
        "WHITE BOX",
    ],
}

PACKAGE_LABELS = {
    "caixa_azul_branca": "cx. azul e branca",
    "caixa_nova_preta": "cx. nova preta",
    "caixa_branca": "cx. branca",
}

AFTERMARKET_TERMS = [
    "COMPATIVEL",
    "COMPATIBLE",
    "REMANUFATURADO",
    "REMANUFACTURED",
    "RECARREGADO",
    "RECARGA",
    "GENERICO",
    "GENERIC",
    "ALTERNATIVO",
]

NOISE_CODES = {
    "TON",
    "TONER",
    "PROD",
    "DRU",
    "CART",
    "CARTUCHO",
    "IMP",
    "HP",
    "LASER",
    "LASERJET",
    "OFFICEJET",
    "COLOR",
    "COLOUR",
    "CROMA",
    "PRETO",
    "CIANO",
    "CYAN",
    "MAGENTA",
    "AMARELO",
    "YELLOW",
    "BLACK",
    "BRANCA",
    "BRANCO",
    "AZUL",
    "NOVA",
    "NOVO",
}


@dataclass(frozen=True)
class Product:
    row_number: int
    sku: str
    name: str
    brand: str | None
    code: str | None
    codes: list[str]
    kind: str | None
    color: str | None
    package: str | None


@dataclass
class Candidate:
    sku: str
    product_name: str
    expected_code: str
    expected_brand: str
    expected_color: str
    expected_kind: str
    expected_package: str
    source: str
    rank: int
    query: str
    title: str
    image_url: str
    page_url: str
    local_path: str = ""
    score: int = 0
    status: str = "revisar"
    reasons: list[str] = field(default_factory=list)
    evidence: str = ""
    ocr_text: str = ""
    visual_tags: str = ""

    def as_row(self, approved: str = "") -> dict[str, Any]:
        return {
            "approved": approved,
            "sku": self.sku,
            "product_name": self.product_name,
            "expected_code": self.expected_code,
            "expected_brand": self.expected_brand,
            "expected_color": self.expected_color,
            "expected_kind": self.expected_kind,
            "expected_package": self.expected_package,
            "source": self.source,
            "rank": self.rank,
            "status": self.status,
            "score": self.score,
            "query": self.query,
            "title": self.title,
            "image_url": self.image_url,
            "page_url": self.page_url,
            "local_path": self.local_path,
            "reasons": " | ".join(self.reasons),
            "ocr_text": self.ocr_text,
            "visual_tags": self.visual_tags,
        }


def main() -> int:
    args = parse_args()
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if args.apply_approvals:
        copied = apply_approvals(args.apply_approvals)
        print(f"Aprovadas copiadas para {APPROVED_RAW_DIR}: {copied} arquivo(s).")
        print("Depois rode: python otimizador_imagens.py --input products/_approved_raw --white-background")
        return 0

    products = load_products(args)
    skipped_existing = getattr(args, "_skipped_existing_products", [])
    if skipped_existing:
        write_skipped_products(skipped_existing, args)
        print(f"SKUs pulados por ja terem imagens prontas/localizadas: {len(skipped_existing)}")
        print(f"Relatorio de pulados: {SKIPPED_PRODUCTS_CSV}")
    else:
        write_skipped_products([], args)
    if not products:
        print("Nenhum produto encontrado.")
        return 0

    print(f"Fonte de SKUs: {args.workbook}")
    if args.local_root:
        print(f"Fonte local de imagens candidatas: {args.local_root}")
    print(f"Produtos selecionados: {len(products)}")
    if args.download:
        print(f"Destino dos downloads: {download_root(args)}\\[SKU]")
    print(f"OCR: {'ativo' if ocr_available() else 'indisponivel'}")

    candidates: list[Candidate] = []
    for index, product in enumerate(products, start=1):
        print(f"[{index}/{len(products)}] {product.sku} - {product.name}")
        product_candidates: list[Candidate] = []

        if args.local_root:
            product_candidates.extend(collect_local_candidates(product, args.local_root, args))

        if args.web:
            product_candidates.extend(collect_duckduckgo_candidates(product, args))

        product_candidates = dedupe_candidates(product_candidates)
        product_candidates.sort(key=lambda item: (item.status_rank(), item.score), reverse=True)  # type: ignore[attr-defined]
        candidates.extend(product_candidates[: args.keep_per_product])

    write_reports(candidates)
    print(f"CSV gerado: {CANDIDATES_CSV}")
    print(f"CSV para aprovacao manual: {APPROVAL_TEMPLATE_CSV}")
    print(f"HTML de revisao: {REVIEW_HTML}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Busca e pontua candidatas de imagem por SKU. "
            "A aprovacao final continua humana quando a evidencia nao for forte."
        )
    )
    parser.add_argument(
        "--source-root",
        type=Path,
        default=DEFAULT_SOURCE_ROOT,
        help=(
            "Pasta onde ficam a planilha e as subpastas de imagens por SKU. "
            "Padrao: Desktop/copia de produtos/Produtos."
        ),
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help=(
            f"Planilha do catalogo. Padrao: [source-root]/{DEFAULT_WORKBOOK_NAME}. "
            "As SKUs sempre saem desta planilha, nao das pastas."
        ),
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--sku", action="append", help="Processa somente um SKU. Pode repetir.")
    parser.add_argument("--limit", type=int, help="Limita a quantidade de produtos.")
    parser.add_argument(
        "--only-missing",
        action="store_true",
        help="Processa apenas linhas sem OK/valor na coluna Imagem.",
    )
    parser.add_argument(
        "--local-root",
        type=Path,
        default=None,
        help=(
            "Pasta com imagens ja baixadas em subpastas [SKU]. "
            "Nao define quais SKUs serao processados; isso vem somente da planilha. "
            "Padrao: [source-root]."
        ),
    )
    parser.add_argument(
        "--no-local-source",
        action="store_true",
        help="Nao usa [source-root]/[SKU] como fonte local de imagens candidatas.",
    )
    parser.add_argument("--web", action="store_true", help="Coleta candidatas no DuckDuckGo Images.")
    parser.add_argument("--download", action="store_true", help="Baixa candidatas web que passaram no filtro de codigo.")
    parser.add_argument(
        "--download-root",
        type=Path,
        default=None,
        help=(
            "Pasta raiz onde as candidatas baixadas serao salvas em [raiz]/[SKU]. "
            "Padrao: [source-root], ou seja, Desktop/copia de produtos/Produtos. "
            f"Para baixar dentro do MSN, use {PRODUCTS_DIR}."
        ),
    )
    parser.add_argument(
        "--include-existing-products",
        action="store_true",
        help="Nao pula SKUs que ja possuem imagens em products/[SKU] ou source-root/[SKU].",
    )
    parser.add_argument("--max-results", type=int, default=8, help="Resultados web analisados por consulta.")
    parser.add_argument("--max-queries", type=int, default=2, help="Consultas web por produto.")
    parser.add_argument("--keep-per-product", type=int, default=8, help="Candidatas mantidas no relatorio por produto.")
    parser.add_argument(
        "--download-per-product",
        type=int,
        default=6,
        help="Quantidade maxima de imagens baixadas por produto.",
    )
    parser.add_argument("--delay", type=float, default=1.2, help="Pausa entre consultas web.")
    parser.add_argument("--apply-approvals", type=Path, help="CSV de aprovacao preenchido com approved=1/sim/ok.")
    args = parser.parse_args()
    normalize_runtime_paths(args)
    return args


def normalize_runtime_paths(args: argparse.Namespace) -> None:
    args.source_root = resolve_source_root(getattr(args, "source_root", DEFAULT_SOURCE_ROOT))
    args.workbook = resolve_workbook_path(args)
    if getattr(args, "local_root", None) is None and not getattr(args, "no_local_source", False):
        args.local_root = args.source_root
    elif getattr(args, "local_root", None) is not None:
        args.local_root = args.local_root.expanduser().resolve()
    if getattr(args, "download_root", None) is None:
        args.download_root = args.source_root
    else:
        args.download_root = args.download_root.expanduser().resolve()


def resolve_source_root(source_root: Path | None) -> Path:
    candidate = (source_root or DEFAULT_SOURCE_ROOT).expanduser()
    if candidate.exists():
        return candidate.resolve()

    discovered = discover_default_source_root()
    if discovered is not None:
        return discovered

    return candidate.resolve()


def discover_default_source_root() -> Path | None:
    desktop = Path.home() / "Desktop"
    if not desktop.exists():
        return None

    for folder in sorted(desktop.iterdir()):
        if not folder.is_dir():
            continue
        normalized = normalize_header(folder.name)
        if "copia" not in normalized or "produtos" not in normalized:
            continue
        product_root = folder / "Produtos"
        workbook = product_root / DEFAULT_WORKBOOK_NAME
        if workbook.exists():
            return product_root.resolve()
    return None


def resolve_workbook_path(args: argparse.Namespace) -> Path:
    workbook_arg = getattr(args, "workbook", None)
    if workbook_arg is not None:
        workbook = workbook_arg.expanduser()
    else:
        source_root = resolve_source_root(getattr(args, "source_root", DEFAULT_SOURCE_ROOT))
        workbook = source_root / DEFAULT_WORKBOOK_NAME

    if workbook.exists():
        return workbook.resolve()

    if workbook.name == DEFAULT_WORKBOOK_NAME:
        discovered = discover_default_source_root()
        if discovered is not None:
            discovered_workbook = discovered / DEFAULT_WORKBOOK_NAME
            if discovered_workbook.exists():
                return discovered_workbook.resolve()

    return workbook.resolve()


def load_products(args: argparse.Namespace) -> list[Product]:
    if not hasattr(args, "source_root"):
        args.source_root = DEFAULT_SOURCE_ROOT
    workbook = resolve_workbook_path(args)
    args.workbook = workbook
    if not workbook.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook}")

    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
    headers = [normalize_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    wanted = {normalize_sku(value) for value in getattr(args, "sku", None) or []}

    products: list[Product] = []
    for row_number in range(2, ws.max_row + 1):
        row = {headers[col - 1]: ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)}
        sku = clean_cell(first_present(row, "sku", "codigo", "codigoproduto"))
        name = clean_cell(first_present(row, "nome", "produto", "descricao"))
        image_status = clean_cell(first_present(row, "imagem", "image"))
        if not sku or not name:
            continue
        if wanted and normalize_sku(sku) not in wanted:
            continue
        if getattr(args, "only_missing", False) and image_status:
            continue

        brand = infer_brand(sku, name)
        codes = infer_codes(sku, name)
        code = codes[0] if codes else None
        kind = infer_kind(sku, name)
        products.append(
            Product(
                row_number=row_number,
                sku=str(sku).strip(),
                name=str(name).strip(),
                brand=brand,
                code=code,
                codes=codes,
                kind=kind,
                color=infer_color(sku, name, kind),
                package=infer_package(sku, name),
            )
        )

    wb.close()
    if not getattr(args, "include_existing_products", False):
        kept_products: list[Product] = []
        skipped_products: list[Product] = []
        for product in products:
            if product_has_existing_images(product, args):
                skipped_products.append(product)
            else:
                kept_products.append(product)
        products = kept_products
        args._skipped_existing_products = skipped_products
    else:
        args._skipped_existing_products = []

    limit = getattr(args, "limit", None)
    if limit is not None:
        products = products[: max(0, limit)]
    return products


def download_root(args: argparse.Namespace) -> Path:
    return args.download_root.expanduser().resolve()


def final_product_dir(product: Product) -> Path:
    return PRODUCTS_DIR / safe_folder_name(product.sku)


def product_has_existing_images(product: Product, args: argparse.Namespace) -> bool:
    return bool(product_existing_locations(product, args))


def product_existing_locations(product: Product, args: argparse.Namespace) -> list[Path]:
    locations: list[Path] = []
    for folder in product_existing_dirs(product, args):
        if folder.exists() and any_product_image(folder):
            locations.append(folder)
    return locations


def product_existing_dirs(product: Product, args: argparse.Namespace) -> list[Path]:
    dirs = [final_product_dir(product)]
    source_root = getattr(args, "source_root", None)
    if source_root is not None:
        dirs.append(Path(source_root) / safe_folder_name(product.sku))

    unique_dirs: list[Path] = []
    seen: set[str] = set()
    for folder in dirs:
        key = str(folder.expanduser().resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        unique_dirs.append(folder)
    return unique_dirs


def any_product_image(root: Path) -> bool:
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        if any(part.startswith("_") for part in path.relative_to(root).parts[:-1]):
            continue
        return True
    return False


def product_image_count(root: Path) -> int:
    if not root.exists():
        return 0
    count = 0
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        if any(part.startswith("_") for part in path.relative_to(root).parts[:-1]):
            continue
        count += 1
    return count


def collect_local_candidates(product: Product, local_root: Path, args: argparse.Namespace) -> list[Candidate]:
    product_dir = local_root.expanduser() / safe_folder_name(product.sku)
    if not product_dir.exists():
        return []

    candidates: list[Candidate] = []
    rank = 1
    for path in sorted(product_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        candidate = Candidate(
            sku=product.sku,
            product_name=product.name,
            expected_code=product.code or "",
            expected_brand=product.brand or "",
            expected_color=product.color or "",
            expected_kind=product.kind or "",
            expected_package=PACKAGE_LABELS.get(product.package or "", product.package or ""),
            source="local",
            rank=rank,
            query="",
            title=path.name,
            image_url=path.resolve().as_uri(),
            page_url="",
            local_path=str(path.resolve()),
        )
        analyze_image_candidate(candidate)
        score_candidate(candidate, product)
        maybe_ocr_candidate(candidate, product, args)
        candidates.append(candidate)
        rank += 1
    return candidates


def collect_duckduckgo_candidates(product: Product, args: argparse.Namespace) -> list[Candidate]:
    candidates: list[Candidate] = []
    seen_urls: set[str] = set()
    queries = build_queries(product)[: args.max_queries]

    for query_index, query in enumerate(queries, start=1):
        try:
            results = duckduckgo_image_search(query, args.max_results)
        except Exception as exc:
            candidates.append(
                Candidate(
                    sku=product.sku,
                    product_name=product.name,
                    expected_code=product.code or "",
                    expected_brand=product.brand or "",
                    expected_color=product.color or "",
                    expected_kind=product.kind or "",
                    expected_package=PACKAGE_LABELS.get(product.package or "", product.package or ""),
                    source="ddg_error",
                    rank=query_index,
                    query=query,
                    title=f"erro na busca: {exc}",
                    image_url="",
                    page_url="",
                    status="erro",
                    reasons=[str(exc)],
                )
            )
            continue

        for rank, item in enumerate(results, start=1):
            image_url = str(item.get("image") or "")
            if not image_url or image_url in seen_urls:
                continue
            seen_urls.add(image_url)
            candidate = Candidate(
                sku=product.sku,
                product_name=product.name,
                expected_code=product.code or "",
                expected_brand=product.brand or "",
                expected_color=product.color or "",
                expected_kind=product.kind or "",
                expected_package=PACKAGE_LABELS.get(product.package or "", product.package or ""),
                source="duckduckgo",
                rank=rank,
                query=query,
                title=str(item.get("title") or ""),
                image_url=image_url,
                page_url=str(item.get("url") or ""),
            )
            score_candidate(candidate, product)
            candidates.append(candidate)

        time.sleep(max(0.0, args.delay))

    candidates = dedupe_candidates(candidates)
    candidates.sort(key=lambda item: (candidate_status_rank(item), item.score), reverse=True)

    if args.download:
        downloaded = 0
        download_limit = max(0, getattr(args, "download_per_product", 6))
        for candidate in candidates:
            if downloaded >= download_limit:
                break
            if candidate.status in {"rejeitada", "erro"}:
                continue
            download_candidate(candidate, product, args)
            if not candidate.local_path:
                continue
            analyze_image_candidate(candidate)
            score_candidate(candidate, product)
            maybe_ocr_candidate(candidate, product, args)
            downloaded += 1
        candidates.sort(key=lambda item: (candidate_status_rank(item), item.score), reverse=True)

    return candidates


def duckduckgo_image_search(query: str, max_results: int) -> list[dict[str, Any]]:
    search_url = "https://duckduckgo.com/?" + urlencode({"q": query, "iax": "images", "ia": "images"})
    html_text = fetch_text(search_url)
    token_match = re.search(r"vqd=([\d-]+)&", html_text) or re.search(r"vqd=['\"]([\d-]+)['\"]", html_text)
    if not token_match:
        raise RuntimeError("token vqd nao encontrado")

    api_url = "https://duckduckgo.com/i.js?" + urlencode(
        {
            "l": "us-en",
            "o": "json",
            "q": query,
            "vqd": token_match.group(1),
            "f": ",,,",
            "p": "1",
        }
    )
    payload = fetch_text(api_url, referer=search_url)
    data = json.loads(payload)
    return list(data.get("results", []))[:max_results]


def fetch_text(url: str, referer: str | None = None) -> str:
    headers = {
        "User-Agent": USER_AGENT,
        "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    }
    if referer:
        headers["Referer"] = referer
    request = Request(url, headers=headers)
    with urlopen(request, timeout=25) as response:
        return response.read().decode("utf-8", errors="replace")


def download_candidate(candidate: Candidate, product: Product, args: argparse.Namespace) -> None:
    target_dir = download_root(args) / safe_folder_name(product.sku)
    target_dir.mkdir(parents=True, exist_ok=True)
    parsed = urlparse(candidate.image_url)
    suffix = Path(unquote(parsed.path)).suffix.lower()
    if suffix not in IMAGE_EXTENSIONS:
        suffix = ".jpg"

    domain = safe_folder_name(parsed.netloc or "imagem")[:40]
    digest = hashlib.sha1(candidate.image_url.encode("utf-8")).hexdigest()[:10]
    target = target_dir / f"{candidate.rank:02d}-{domain}-{digest}{suffix}"

    if target.exists():
        candidate.local_path = str(target.resolve())
        return

    try:
        request = Request(candidate.image_url, headers={"User-Agent": USER_AGENT})
        with urlopen(request, timeout=35) as response:
            content_type = response.headers.get("Content-Type", "")
            guessed = mimetypes.guess_extension(content_type.split(";")[0].strip())
            if guessed and guessed.lower() in IMAGE_EXTENSIONS and target.suffix == ".jpg":
                target = target.with_suffix(guessed.lower())
            data = response.read()
        target.write_bytes(data)
        candidate.local_path = str(target.resolve())
    except (HTTPError, URLError, TimeoutError, OSError) as exc:
        candidate.reasons.append(f"download falhou: {exc}")


def score_candidate(candidate: Candidate, product: Product) -> None:
    evidence_text = " ".join(
        [
            candidate.title,
            "" if candidate.image_url.startswith("file:///") else candidate.image_url,
            candidate.page_url,
            Path(candidate.local_path).name if candidate.local_path else "",
            candidate.ocr_text,
        ]
    )
    candidate.evidence = " ".join([evidence_text, candidate.visual_tags]).strip()
    normalized_words = normalize_words(evidence_text)
    normalized_compact = normalize_compact(evidence_text)

    score = 0
    reasons: list[str] = []

    code_match, matched_code, exact_code_match = find_code_match(normalized_compact, product)
    if code_match:
        score += 70 if exact_code_match else 58
        if exact_code_match:
            reasons.append(f"codigo {product.code} encontrado")
        else:
            reasons.append(f"variante-base {matched_code} encontrada para {product.code}")
    else:
        score -= 60
        reasons.append(f"codigo {product.code or 'desconhecido'} nao encontrado")

    conflict_codes = find_conflicting_codes(evidence_text, product)
    if conflict_codes:
        score -= 120
        reasons.append("codigo conflitante: " + ", ".join(conflict_codes[:5]))

    conflicting_colors = find_conflicting_colors(normalized_words, product.color)
    if conflicting_colors:
        score -= 70
        reasons.append("cor conflitante: " + ", ".join(conflicting_colors))

    if product.brand:
        brand_match = text_has_any(normalized_words, BRAND_ALIASES.get(product.brand, [product.brand.upper()]))
        official = is_official_or_trusted(candidate.image_url, candidate.page_url, product.brand)
        if brand_match:
            score += 20
            reasons.append(f"marca {product.brand} encontrada")
        if official:
            score += 25
            reasons.append(f"fonte/imagem oficial ou confiavel para {product.brand}")
        if not brand_match and not official:
            score -= 15
            reasons.append(f"marca {product.brand} nao apareceu")
    else:
        official = False

    aftermarket_match = text_has_any(normalized_words, AFTERMARKET_TERMS)
    product_allows_aftermarket = text_has_any(normalize_words(product.name), AFTERMARKET_TERMS)
    if aftermarket_match and not product_allows_aftermarket:
        score -= 90
        reasons.append("termo de compativel/remanufaturado sem isso constar na planilha")

    if product.kind:
        if text_has_any(normalized_words, KIND_SYNONYMS.get(product.kind, [product.kind.upper()])):
            score += 10
            reasons.append(f"tipo {product.kind} encontrado")
        else:
            reasons.append(f"tipo {product.kind} nao confirmado")

    if product.color:
        if text_has_any(normalized_words, COLOR_SYNONYMS.get(product.color, [product.color.upper()])):
            score += 12
            reasons.append(f"cor {product.color} encontrada")
        else:
            score -= 8
            reasons.append(f"cor {product.color} nao confirmada")

    package_confirmed = False
    package_conflict = False
    if product.package:
        package_label = PACKAGE_LABELS.get(product.package, product.package)
        if package_matches(normalized_words, candidate.visual_tags, product.package):
            score += 18
            package_confirmed = True
            reasons.append(f"embalagem {package_label} confirmada")
        else:
            conflicts = find_conflicting_packages(normalized_words, product.package)
            if conflicts:
                score -= 55
                package_conflict = True
                reasons.append("embalagem conflitante: " + ", ".join(conflicts))
            else:
                score -= 12
                reasons.append(f"embalagem {package_label} nao confirmada")

    if candidate.source == "local":
        score += 5
        reasons.append("imagem ja esta na pasta local do SKU")

    if conflict_codes:
        status = "rejeitada"
    elif conflicting_colors:
        status = "rejeitada"
    elif package_conflict:
        status = "rejeitada"
    elif aftermarket_match and not product_allows_aftermarket:
        status = "rejeitada"
    elif not code_match:
        status = "rejeitada"
    elif score >= 105 and official:
        status = "forte"
    elif score >= 80:
        status = "boa"
    else:
        status = "revisar"

    if product.package and not package_confirmed and status in {"forte", "boa"}:
        status = "revisar"
        reasons.append("embalagem exigida pela planilha; manter em revisao ate confirmar visualmente")

    candidate.score = score
    candidate.status = status
    candidate.reasons = reasons


def maybe_ocr_candidate(candidate: Candidate, product: Product, args: argparse.Namespace) -> None:
    if not candidate.local_path or not ocr_available():
        return
    text = run_ocr(Path(candidate.local_path))
    if not text:
        return
    candidate.ocr_text = text
    score_candidate(candidate, product)
    candidate.reasons.append("OCR analisado")


def analyze_image_candidate(candidate: Candidate) -> None:
    if not candidate.local_path:
        return
    tags = detect_visual_tags(Path(candidate.local_path))
    if tags:
        candidate.visual_tags = " ".join(tags)


def detect_visual_tags(path: Path) -> list[str]:
    try:
        from PIL import Image, ImageOps

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            with Image.open(path) as original:
                image = ImageOps.exif_transpose(original).convert("RGBA")
        image.thumbnail((240, 240))
        if hasattr(image, "get_flattened_data"):
            pixels = list(image.get_flattened_data())
        else:
            pixels = list(image.getdata())
    except Exception:
        return []

    counts = {
        "VISUAL_BLUE": 0,
        "VISUAL_DARK": 0,
        "VISUAL_CYAN": 0,
        "VISUAL_MAGENTA": 0,
        "VISUAL_YELLOW": 0,
        "VISUAL_RED": 0,
    }
    total = 0
    for red, green, blue, alpha in pixels:
        if alpha < 32:
            continue
        total += 1
        if red > 245 and green > 245 and blue > 245:
            continue
        if max(red, green, blue) < 75:
            counts["VISUAL_DARK"] += 1
        if blue > 125 and blue > red * 1.25 and blue > green * 1.08:
            counts["VISUAL_BLUE"] += 1
        if green > 110 and blue > 110 and red < 105 and abs(green - blue) < 80:
            counts["VISUAL_CYAN"] += 1
        if red > 135 and blue > 105 and green < 115:
            counts["VISUAL_MAGENTA"] += 1
        if red > 145 and green > 125 and blue < 105:
            counts["VISUAL_YELLOW"] += 1
        if red > 140 and green < 95 and blue < 95:
            counts["VISUAL_RED"] += 1

    if total == 0:
        return []

    tags: list[str] = []
    for tag, count in counts.items():
        ratio = count / total
        if ratio >= 0.015:
            tags.append(tag)
    return tags


def ocr_available() -> bool:
    try:
        import pytesseract  # noqa: F401
        from PIL import Image  # noqa: F401
    except Exception:
        return False
    return find_tesseract_cmd() is not None and bool(ocr_language_string())


def run_ocr(path: Path) -> str:
    try:
        import pytesseract
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps

        tesseract_cmd = find_tesseract_cmd()
        if tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = str(tesseract_cmd)

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            with Image.open(path) as image:
                normalized = ImageOps.exif_transpose(image)
            normalized = normalized.convert("L")
            if normalized.width < 1400:
                scale = min(3, max(2, round(1400 / max(1, normalized.width))))
                normalized = normalized.resize(
                    (normalized.width * scale, normalized.height * scale),
                    Image.Resampling.LANCZOS,
                )
            normalized = ImageOps.autocontrast(normalized)
            normalized = ImageEnhance.Contrast(normalized).enhance(1.7)
            normalized = normalized.filter(ImageFilter.SHARPEN)
            tessdata_dir = find_tessdata_dir()
            config_parts = ["--psm 6"]
            if tessdata_dir:
                os.environ["TESSDATA_PREFIX"] = str(tessdata_dir)
            text = pytesseract.image_to_string(
                normalized,
                lang=ocr_language_string(),
                config=" ".join(config_parts),
            )
        return re.sub(r"\s+", " ", text).strip()
    except Exception:
        return ""


def find_tesseract_cmd() -> Path | None:
    from_path = shutil.which("tesseract")
    candidates = [
        Path(from_path) if from_path else None,
        Path.home() / "AppData" / "Local" / "Programs" / "Tesseract-OCR" / "tesseract.exe",
        Path("C:/Program Files/Tesseract-OCR/tesseract.exe"),
        Path("C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"),
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def find_tessdata_dir() -> Path | None:
    candidates = [
        Path.home() / "AppData" / "Local" / "Tesseract-OCR" / "tessdata",
        Path("C:/Program Files/Tesseract-OCR/tessdata"),
        Path("C:/Program Files (x86)/Tesseract-OCR/tessdata"),
    ]
    for candidate in candidates:
        if candidate.exists() and any(candidate.glob("*.traineddata")):
            return candidate
    return None


def ocr_language_string() -> str:
    tessdata_dir = find_tessdata_dir()
    if not tessdata_dir:
        return ""
    languages = {path.stem for path in tessdata_dir.glob("*.traineddata")}
    selected = [lang for lang in ("eng", "por") if lang in languages]
    return "+".join(selected)


def find_conflicting_codes(text: str, product: Product) -> list[str]:
    expected = normalize_compact(product.code or "")
    accepted = {compact for _, compact, _ in acceptable_code_variants(product)}
    if not expected:
        return []
    expected_prefix = re.match(r"^[A-Z]+", expected)
    prefix = expected_prefix.group(0) if expected_prefix else ""
    conflicts: list[str] = []
    for raw in infer_codes("", text):
        code = normalize_compact(raw)
        if not code or code in accepted:
            continue
        raw_prefix_match = re.match(r"^[A-Z]+", code)
        raw_prefix = raw_prefix_match.group(0) if raw_prefix_match else ""
        if prefix and raw_prefix == prefix and abs(len(code) - len(expected)) <= 2:
            conflicts.append(raw)
    return unique_list(conflicts)


def find_conflicting_colors(normalized_words: str, expected_color: str | None) -> list[str]:
    if not expected_color:
        return []
    conflicts: list[str] = []
    for color, synonyms in COLOR_SYNONYMS.items():
        if color == expected_color:
            continue
        if text_has_any(normalized_words, synonyms):
            conflicts.append(color)
    return conflicts


def package_matches(normalized_words: str, visual_tags: str, expected_package: str) -> bool:
    if text_has_any(normalized_words, PACKAGE_SYNONYMS.get(expected_package, [])):
        return True
    visual_words = normalize_words(visual_tags)
    if expected_package == "caixa_azul_branca":
        return text_has_any(visual_words, ["VISUAL BLUE"])
    return False


def find_conflicting_packages(normalized_words: str, expected_package: str) -> list[str]:
    conflicts: list[str] = []
    for package, synonyms in PACKAGE_SYNONYMS.items():
        if package == expected_package:
            continue
        if text_has_any(normalized_words, synonyms):
            conflicts.append(PACKAGE_LABELS.get(package, package))
    return conflicts


def find_code_match(normalized_compact_text: str, product: Product) -> tuple[bool, str, bool]:
    for raw, compact, exact in acceptable_code_variants(product):
        if compact and compact in normalized_compact_text:
            return True, raw, exact
    return False, "", False


def acceptable_code_variants(product: Product) -> list[tuple[str, str, bool]]:
    variants: list[tuple[str, str, bool]] = []

    def add(raw: str, exact: bool) -> None:
        compact = normalize_compact(raw)
        if compact and all(existing[1] != compact for existing in variants):
            variants.append((raw, compact, exact))

    for code in product.codes or ([product.code] if product.code else []):
        if not code:
            continue
        add(code, True)
        compact = normalize_compact(code)
        base = strip_hp_regional_suffix(compact)
        if base and base != compact:
            add(base, False)
    return variants


def strip_hp_regional_suffix(compact_code: str) -> str | None:
    # HP often appends regional/contract suffixes to a base cartridge code.
    # Examples seen in this catalog: CE313AB -> CE313A, CF322AC -> CF322A,
    # CF320XC -> CF320X, Q5942YC -> Q5942Y.
    match = re.match(r"^([A-Z]{1,4}\d{3,5}[A-Z])([A-Z])$", compact_code)
    if match and match.group(2) in {"B", "C"}:
        return match.group(1)
    return None


def is_official_or_trusted(image_url: str, page_url: str, brand: str) -> bool:
    domains = BRAND_DOMAINS.get(brand, [])
    text = " ".join([urlparse(image_url).netloc.lower(), urlparse(page_url).netloc.lower()])
    return any(domain.lower() in text for domain in domains)


def text_has_any(normalized_words: str, values: list[str]) -> bool:
    padded = f" {normalized_words} "
    for value in values:
        target = normalize_words(value)
        if f" {target} " in padded:
            return True
    return False


def build_queries(product: Product) -> list[str]:
    brand = product.brand or ""
    code = product.code or ""
    kind = KIND_LABELS.get(product.kind or "", product.kind or "")
    color = product.color or ""
    package_label = PACKAGE_LABELS.get(product.package or "", "")
    queries = [
        product.name,
        " ".join(part for part in [brand, code, kind, color, package_label, "original"] if part),
        " ".join(part for part in [brand, code, "cartridge image"] if part),
        " ".join(part for part in [brand, code, package_label, "image"] if part),
        f'"{brand} {code}"' if brand and code else code,
        product.sku,
    ]
    return unique_list([query.strip() for query in queries if query.strip()])


def infer_brand(sku: str, name: str) -> str | None:
    text = normalize_words(f"{sku} {name}")
    for brand, aliases in BRAND_ALIASES.items():
        if text_has_any(text, aliases):
            return brand
    return None


def infer_kind(sku: str, name: str) -> str | None:
    text = normalize_words(f"{sku} {name}")
    if text.startswith("TON ") or text_has_any(text, KIND_SYNONYMS["toner"]):
        return "toner"
    if text.startswith("DRU ") or text_has_any(text, KIND_SYNONYMS["drum"]):
        return "drum"
    if text.startswith("IMP ") or text_has_any(text, KIND_SYNONYMS["impressora"]):
        return "impressora"
    for kind, synonyms in KIND_SYNONYMS.items():
        if text_has_any(text, synonyms):
            return kind
    return None


def infer_color(sku: str, name: str, kind: str | None = None) -> str | None:
    text = normalize_words(name)
    for color, synonyms in COLOR_SYNONYMS.items():
        if text_has_any(text, synonyms):
            return color

    color_suffix_allowed = kind in {"toner", "cartucho", "developer", "tinta", "cera"}
    if not color_suffix_allowed:
        color_suffix_allowed = text_has_any(text, ["DEVELOPER", "REVELADOR", "TINTA", "INK", "CERA", "COLORQUBE"])
    if not color_suffix_allowed:
        return None

    parts = re.split(r"[-_\s/]+", normalize_words(sku))
    for part in reversed(parts):
        color = COLOR_SUFFIXES.get(part)
        if color:
            return color
    return None


def infer_package(sku: str, name: str) -> str | None:
    text = normalize_words(name)
    # Order matters: "azul e branca" contains "branca".
    for package in ("caixa_azul_branca", "caixa_nova_preta", "caixa_branca"):
        if text_has_any(text, PACKAGE_SYNONYMS[package]):
            return package
    sku_parts = re.split(r"[-_\s/]+", normalize_words(sku))
    if "AZ" in sku_parts:
        return "caixa_azul_branca"
    if "NV" in sku_parts:
        return "caixa_nova_preta"
    if "BR" in sku_parts:
        return "caixa_branca"
    return None


def infer_codes(sku: str, name: str) -> list[str]:
    sources = [(sku, True), (name, False)]
    patterns = [
        r"\b[A-Z]{1,8}-?[A-Z]?\d[A-Z0-9-]{1,}\b",
        r"\b\d[A-Z0-9-]{3,}\b",
    ]
    codes: list[str] = []
    for source, allow_short in sources:
        text = normalize_words(source)
        for pattern in patterns:
            for match in re.findall(pattern, text):
                cleaned = clean_code(match, allow_short=allow_short)
                if cleaned:
                    codes.append(cleaned)

        if allow_short:
            for piece in re.split(r"[-_\s/]+", text):
                cleaned = clean_code(piece, allow_short=True)
                if cleaned:
                    codes.append(cleaned)
    return unique_list(codes)


def clean_code(value: str, allow_short: bool = False) -> str | None:
    code = normalize_words(value).replace(" ", "")
    code = code.strip("-_./")
    compact = normalize_compact(code)
    if not compact or compact in NOISE_CODES:
        return None
    if compact in COLOR_SUFFIXES:
        return None
    if not any(char.isdigit() for char in compact):
        return None
    if len(compact) < (3 if allow_short else 4):
        return None
    return code


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9_]", "", strip_accents(str(value or "")).lower())


def normalize_words(value: Any) -> str:
    text = strip_accents(unquote(str(value or ""))).upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_compact(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", normalize_words(value))


def normalize_sku(value: str) -> str:
    return safe_folder_name(value).upper()


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_folder_name(value: str) -> str:
    text = strip_accents(str(value)).strip()
    text = re.sub(r"[<>:\"/\\|?*\x00-\x1F]", "-", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .") or "produto"


def unique_list(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        key = normalize_compact(value)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def dedupe_candidates(candidates: list[Candidate]) -> list[Candidate]:
    result: list[Candidate] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = candidate.local_path or candidate.image_url or candidate.page_url or candidate.title
        digest = hashlib.sha1(key.encode("utf-8", errors="ignore")).hexdigest()
        if digest in seen:
            continue
        seen.add(digest)
        result.append(candidate)
    return result


def candidate_status_rank(candidate: Candidate) -> int:
    return {"forte": 4, "boa": 3, "revisar": 2, "erro": 1, "rejeitada": 0}.get(candidate.status, 0)


Candidate.status_rank = candidate_status_rank  # type: ignore[attr-defined]


def write_reports(candidates: list[Candidate]) -> None:
    fields = list(
        Candidate(
            sku="",
            product_name="",
            expected_code="",
            expected_brand="",
            expected_color="",
            expected_kind="",
            expected_package="",
            source="",
            rank=0,
            query="",
            title="",
            image_url="",
            page_url="",
        ).as_row().keys()
    )
    with CANDIDATES_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for candidate in candidates:
            writer.writerow(candidate.as_row())

    with APPROVAL_TEMPLATE_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for candidate in candidates:
            approved = "1" if candidate.status == "forte" else ""
            writer.writerow(candidate.as_row(approved=approved))

    REVIEW_HTML.write_text(render_review_html(candidates), encoding="utf-8")


def write_skipped_products(products: list[Product], args: argparse.Namespace) -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    fields = ["sku", "name", "existing_locations", "image_count"]
    with SKIPPED_PRODUCTS_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for product in products:
            locations = product_existing_locations(product, args)
            writer.writerow(
                {
                    "sku": product.sku,
                    "name": product.name,
                    "existing_locations": " | ".join(str(location) for location in locations),
                    "image_count": sum(product_image_count(location) for location in locations),
                }
            )


def render_review_html(candidates: list[Candidate]) -> str:
    grouped: dict[str, list[Candidate]] = {}
    for candidate in candidates:
        grouped.setdefault(candidate.sku, []).append(candidate)

    rows = []
    for sku, items in grouped.items():
        product_name = items[0].product_name if items else sku
        cards = "\n".join(render_candidate_card(item) for item in items)
        rows.append(
            f"""
            <section class="product">
              <header>
                <h2>{html.escape(sku)}</h2>
                <p>{html.escape(product_name)}</p>
              </header>
              <div class="cards">{cards}</div>
            </section>
            """
        )

    return f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Revisao de candidatas - MSN</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f4f6f8;
      --ink: #182230;
      --muted: #667085;
      --line: #d9e0e8;
      --good: #107c41;
      --ok: #1769aa;
      --warn: #a15c00;
      --bad: #b42318;
    }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--ink);
      background: var(--bg);
    }}
    .top {{
      position: sticky;
      top: 0;
      z-index: 2;
      padding: 16px 22px;
      background: #ffffff;
      border-bottom: 1px solid var(--line);
    }}
    .top h1 {{
      margin: 0 0 4px;
      font-size: 20px;
    }}
    .top p {{
      margin: 0;
      color: var(--muted);
      font-size: 13px;
    }}
    .product {{
      padding: 18px 22px 24px;
      border-bottom: 1px solid var(--line);
    }}
    .product header {{
      margin-bottom: 12px;
    }}
    .product h2 {{
      margin: 0 0 3px;
      font-size: 18px;
    }}
    .product p {{
      margin: 0;
      color: var(--muted);
    }}
    .cards {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
      gap: 12px;
    }}
    .card {{
      background: #ffffff;
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
    }}
    .card.forte {{ border-color: #8fcca9; }}
    .card.boa {{ border-color: #9fc3e6; }}
    .card.revisar {{ border-color: #e5c07b; }}
    .card.rejeitada {{ opacity: .58; }}
    .thumb {{
      display: block;
      width: 100%;
      aspect-ratio: 1 / 1;
      object-fit: contain;
      background: #fff;
      border-bottom: 1px solid var(--line);
    }}
    .body {{
      padding: 10px;
    }}
    .badge {{
      display: inline-block;
      margin-right: 6px;
      padding: 3px 7px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
      color: #fff;
      background: var(--muted);
    }}
    .forte .badge {{ background: var(--good); }}
    .boa .badge {{ background: var(--ok); }}
    .revisar .badge {{ background: var(--warn); }}
    .rejeitada .badge {{ background: var(--bad); }}
    .title {{
      margin: 9px 0;
      min-height: 38px;
      font-size: 14px;
      line-height: 1.35;
    }}
    .meta, .reasons {{
      margin: 0;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.4;
      overflow-wrap: anywhere;
    }}
    .links {{
      display: flex;
      gap: 10px;
      margin-top: 10px;
      font-size: 13px;
    }}
    a {{ color: #0b5cab; }}
  </style>
</head>
<body>
  <div class="top">
    <h1>Revisao de candidatas de imagens</h1>
    <p>Use o CSV candidatas-aprovacao.csv para marcar approved=1 nas imagens escolhidas. Forte ja vem pre-marcada, mas confira antes de aplicar.</p>
  </div>
  {''.join(rows)}
</body>
</html>
"""


def render_candidate_card(candidate: Candidate) -> str:
    image_src = Path(candidate.local_path).resolve().as_uri() if candidate.local_path else candidate.image_url
    title = candidate.title or Path(candidate.local_path).name or candidate.image_url
    page_link = (
        f'<a href="{html.escape(candidate.page_url)}" target="_blank" rel="noopener">fonte</a>'
        if candidate.page_url
        else ""
    )
    image_link = (
        f'<a href="{html.escape(candidate.image_url)}" target="_blank" rel="noopener">imagem</a>'
        if candidate.image_url
        else ""
    )
    local = html.escape(candidate.local_path)
    return f"""
      <article class="card {html.escape(candidate.status)}">
        {'<img class="thumb" src="' + html.escape(image_src) + '" alt="">' if image_src else ''}
        <div class="body">
          <span class="badge">{html.escape(candidate.status)}</span>
          <span class="meta">score {candidate.score} | {html.escape(candidate.source)} #{candidate.rank}</span>
          <div class="title">{html.escape(title)}</div>
          <p class="reasons">{html.escape(' | '.join(candidate.reasons))}</p>
          <p class="meta">{local}</p>
          <div class="links">{page_link}{image_link}</div>
        </div>
      </article>
    """


def apply_approvals(path: Path) -> int:
    copied = 0
    with path.expanduser().open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            approved = normalize_words(row.get("approved", ""))
            if approved not in {"1", "SIM", "S", "OK", "TRUE", "YES", "Y"}:
                continue
            sku = safe_folder_name(row.get("sku", ""))
            local_path = row.get("local_path", "")
            if not sku or not local_path:
                continue
            source = Path(local_path)
            if not source.exists() or not source.is_file():
                continue
            target_dir = APPROVED_RAW_DIR / sku
            target_dir.mkdir(parents=True, exist_ok=True)
            target = target_dir / source.name
            if target.exists():
                target = target_dir / f"{source.stem}-{hashlib.sha1(str(source).encode()).hexdigest()[:8]}{source.suffix}"
            shutil.copy2(source, target)
            copied += 1
    return copied


if __name__ == "__main__":
    raise SystemExit(main())
